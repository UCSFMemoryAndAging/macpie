"""
Public testing utility functions for Excel files.
"""

import openpyxl as pyxl

from macpie.tools import openpyxltools


def assert_excels_equal(filepath_1, filepath_2):
    """
    For testing equality of :class:`openpyxl.workbook.workbook.Workbook` objects

    Parameters
    ----------
    filepath_1 : Path
        Path of left Excel file to compare
    filepath_2 : Path
        Path of right Excel file to compare
    """
    wb1 = pyxl.load_workbook(filepath_1)
    wb2 = pyxl.load_workbook(filepath_2)

    # same sheets?
    assert set(wb1.sheetnames) == set(wb2.sheetnames)

    for sheet in wb1.sheetnames:
        assert_excel_worksheet_equal(wb1[sheet], wb2[sheet])


def assert_excel_worksheet_equal(left_ws, right_ws):
    """
    For testing equality of :class:`openpyxl.worksheet.worksheet.Worksheet` objects

    Parameters
    ----------
    left_ws : Worksheet
    right_ws : Worksheet
    """

    # same range of data?
    assert left_ws.max_column == right_ws.max_column
    assert left_ws.max_row == right_ws.max_row

    # same data in each cell?
    for left_cell in openpyxltools.iter_cells(left_ws):
        right_cell = right_ws[left_cell.coordinate]
        # convert None to empty string
        left_cell.value = "" if left_cell.value is None else left_cell.value
        right_cell.value = "" if right_cell.value is None else right_cell.value
        assert left_cell.value == right_cell.value, (
            f"'{left_ws.title}'.{left_cell.coordinate} [{repr(left_cell.value)}] != "
            f"'{right_ws.title}'.{right_cell.coordinate} [{repr(right_cell.value)}]"
        )
