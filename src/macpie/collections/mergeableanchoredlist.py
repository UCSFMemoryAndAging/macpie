import warnings

import numpy as np
import pandas as pd
import openpyxl as pyxl

from macpie._config import get_option
from macpie.core.api import Dataset, DatasetFields
from macpie.tools import strtools
from macpie.exceptions import MergeableAnchoredListError, ParserError
from macpie.util import DictLikeDataset

from .anchoredlist import AnchoredList
from .basiclist import BasicList


class MergeableAnchoredList(AnchoredList):
    """A :class:`macpie.AnchoredList` that is `mergeable`,
    meaning all ``secondary`` Datasets have a common column that can be
    used to merge with a column in the ``primary`` Dataset.
    Includes advanced duplicate handling functionality.

    :param primary: The primary `anchor` Dataset of the collection.
    :param secondary: The secondary Datasets of the collection.
    :param primary_anchor_col: The column in ``primary`` to merge on
    :param secondary_anchor_col: The column in each ``secondary`` Dataset
                                 to merge on
    :param selected_fields: Fields to keep (discarding the rest)
    """

    #: Tag that denotes a Dataset can be merged (i.e. no duplicates)
    tag_mergeable = get_option("dataset.tag.mergeable")

    #: Tag that denotes a Dataset has been merged (also means mergeable)
    tag_merged = get_option("dataset.tag.merged")

    #: Tag that denotes a Dataset has not been merged (though mergeable)
    tag_not_merged = get_option("dataset.tag.not_merged")

    #: Tag that denotes a Dataset cannot be merged (i.e. has duplicates)
    tag_duplicates = get_option("dataset.tag.duplicates")

    def __init__(
        self,
        primary: Dataset = None,
        secondary: BasicList = None,
        primary_anchor_col=None,
        secondary_anchor_col=None,
        selected_fields: DatasetFields = None,
    ):
        self._merged_dset = None

        self._primary_anchor_col = primary_anchor_col
        self._secondary_anchor_col = secondary_anchor_col
        self._selected_fields = selected_fields

        super().__init__(primary, secondary)

    def __repr__(self) -> str:
        return (
            f"{self.__class__.__name__}("
            f"primary={self._primary!r}, "
            f"secondary={self._secondary!r}, "
            f"primary_anchor_col={self._primary_anchor_col!r}, "
            f"secondary_anchor_col={self._secondary_anchor_col!r}, "
            f"selected_fields={self._selected_fields!r})"
        )

    @AnchoredList.primary.setter
    def primary(self, dset: Dataset):
        """Sets the `primary` Dataset of this collection."""
        AnchoredList.primary.fset(self, dset)
        if self._primary is not None:
            if self._primary_anchor_col is None:
                self._primary_anchor_col = self._primary.id_col_name
            if self._secondary_anchor_col is None:
                self._secondary_anchor_col = (
                    self._primary.id_col_name + get_option("operators.binary.column_suffixes")[0]
                )

    @property
    def merged_dset(self):
        """The merged data.

        :return: class:`pandas.DataFrame`
        """
        return self._merged_dset

    @property
    def key_fields(self):
        """A list of all :attr:`macpie.Dataset.key_fields` contained
        in this :class:`MergeableAnchoredList`.
        """
        key_fields = super().key_fields
        prim_anchor_field = (self._primary.name, self._primary_anchor_col)
        if self._primary_anchor_col and prim_anchor_field not in key_fields:
            key_fields.append(prim_anchor_field)
        if self._secondary_anchor_col:
            for sec in self._secondary:
                sec_anchor_field = (sec.name, self._secondary_anchor_col)
                if sec_anchor_field not in key_fields:
                    key_fields.append(sec_anchor_field)
        return key_fields

    @property
    def all_fields(self):
        """A list of all :attr:`macpie.Dataset.all_fields` contained
        in this :class:`MergeableAnchoredList`.
        """
        if not self._secondary_anchor_col:
            return super().all_fields

        fields = []
        if self._primary:
            fields.extend(self._primary.all_fields)
        if self._secondary:
            for sec in self._secondary:
                sec_anchor_field = (sec.name, self._secondary_anchor_col)
                fields.append(sec_anchor_field)
                fields.extend(sec.all_fields)
        return fields

    def add_secondary(self, dset: Dataset):
        """Append `dset` to :attr:`MergeableAnchoredList.secondary`."""
        if self._secondary_anchor_col not in dset.columns:
            warnings.warn(
                f"Warning: Secondary dataset '{dset!r}' does not have "
                f"anchor column '{self._secondary_anchor_col}'. Skipping..."
            )
            return

        dups = dset.duplicated(subset=[self._secondary_anchor_col], keep=False)
        if dups.any():
            dups_col = get_option("column.system.duplicates")
            if dups_col in dset.sys_cols:
                dset.rename_col(dups_col, dups_col + "_prior")
            dset.mac.insert(dups_col, dups)
            dset.add_tag(MergeableAnchoredList.tag_duplicates)
        else:
            dset.add_tag(MergeableAnchoredList.tag_mergeable)

        super().add_secondary(dset)

        dset.display_name_generator = MergeableAnchoredList.dataset_display_name_generator

        self._merged_dset = None

    def keep_fields(self, selected_fields, keep_unselected: bool = False):
        """Keep specified fields (and drop the rest).

        :param selected_fields: Fields to keep
        :param keep_unselected: If True, if a Dataset is not in ``selected_fields``,
                                then keep entire Dataset. Defaults to False.
        """
        fields_to_include = []
        selected_dsets = selected_fields.datasets

        if self._primary.name not in selected_dsets:
            fields_to_include.extend(
                field for field in self.key_fields if field[0] == self._primary.name
            )

        fields_to_include.extend(
            [field for field in self.key_fields if field[0] in selected_dsets]
        )
        fields_to_include.extend(
            [field for field in self.sys_fields if field[0] in selected_dsets]
        )

        selected_fields.prepend(fields_to_include)
        selected_fields.sort(self)

        super().keep_fields(selected_fields, keep_unselected=keep_unselected)

    def merge(self):
        """Perform the merge."""
        if self._selected_fields:
            self.keep_fields(self._selected_fields, keep_unselected=True)

        merged_dset = self._primary.copy(deep=True)
        merged_dset.columns = pd.MultiIndex.from_product(
            [[self._primary.name], merged_dset.columns]
        )

        mergeable_secondary = self._secondary.filter(MergeableAnchoredList.tag_mergeable)

        for sec in mergeable_secondary:
            if self._selected_fields and sec.name not in self._selected_fields.datasets:
                sec.add_tag(MergeableAnchoredList.tag_not_merged)
            else:
                merged_dset = merged_dset.mac.merge(
                    sec.copy(deep=True),
                    left_on=[(self._primary.name, self._primary_anchor_col)],
                    right_on=[self._secondary_anchor_col],
                    add_indexes=(None, sec.name),
                )
                sec.add_tag(MergeableAnchoredList.tag_merged)

        # reset index to start from 1 for user readability
        start_index = 1
        merged_dset.index = np.arange(start_index, len(merged_dset) + start_index)

        self._merged_dset = merged_dset

    def get_available_fields(self):
        """Get all "available" fields in this collection.

        :return: :class:`macpie.util.DatasetFields`
        """
        available_fields = DatasetFields.from_collection(
            self,
            tags=[
                MergeableAnchoredList.tag_anchor,
                MergeableAnchoredList.tag_mergeable,
                MergeableAnchoredList.tag_duplicates,
            ],
        )
        available_fields = available_fields.filter(DatasetFields.tag_non_key_field)
        available_fields.title = self._sheetname_available_fields

        available_fields.append_col_fill(None, header=get_option("column.to_merge"))
        print("avaol", available_fields)
        print("avaol title", available_fields.title, self._sheetname_available_fields)
        return available_fields

    def get_duplicates(self):
        dup_dsets = self._secondary.filter(MergeableAnchoredList.tag_duplicates)

        if not dup_dsets:
            return {}

        result = {}

        for dset in dup_dsets:
            result[dset.name] = []

        for dset in dup_dsets:
            dup_rows_df = dset.duplicated(subset=[self._secondary_anchor_col], keep=False)
            result[dset.name].extend(list(dset[dup_rows_df].array))

        return result

    def to_dict(self):
        """Convert the MergeableAnchoredList to a dictionary."""
        return {
            "primary": self._primary,
            "secondary": self._secondary,
            "primary_anchor_col": self._primary_anchor_col,
            "secondary_anchor_col": self._secondary_anchor_col,
            "selected_fields": self._selected_fields,
        }

    def to_excel(self, excel_writer, write_repr=True, merge: bool = True, **kwargs):
        """Write :class:`MergeableAnchoredList` to an Excel file.

        :param merge: If True, output merged result. If False, keep everything
                      unmerged. Defaults to True.
        """

        if merge:
            if not self._merged_dset:
                self.merge()

            self._merged_dset.name = get_option("excel.sheet_name.merged_results")
            self._merged_dset.to_excel(excel_writer, **kwargs)
            self._secondary.filter([MergeableAnchoredList.tag_not_merged]).to_excel(
                excel_writer, write_repr=False, **kwargs
            )
        else:
            self._primary.to_excel(excel_writer, **kwargs)
            self._secondary.filter([MergeableAnchoredList.tag_mergeable]).to_excel(
                excel_writer, write_repr=False, **kwargs
            )

        self._secondary.filter(MergeableAnchoredList.tag_duplicates).to_excel(
            excel_writer, write_repr=False, **kwargs
        )

        self.get_available_fields().to_excel(excel_writer)

        if write_repr:
            self.get_excel_repr().to_excel(excel_writer)

    @staticmethod
    def dataset_display_name_generator(dset: Dataset):
        suffixes = []
        if dset.tags:
            if MergeableAnchoredList.tag_mergeable in dset.tags:
                suffixes = ["linked"]
            elif MergeableAnchoredList.tag_duplicates in dset.tags:
                suffixes = ["DUPS"]
        return strtools.add_suffixes_with_base(dset.name, suffixes, max_length=-1, delimiter="_")

    @classmethod
    def from_excel_dict(cls, filepath, excel_dict) -> "MergeableAnchoredList":
        """Construct :class:`MergeableAnchoredList` from an Excel file."""

        available_fields = DatasetFields.from_excel_sheet_create_tags(
            filepath, sheet_name=get_option("excel.sheet_name.available_fields")
        )

        selected_fields = available_fields.filter(get_option("column.to_merge"))
        selected_fields.title = get_option("excel.sheet_name.selected_fields")

        _primary = None
        _secondary = BasicList()
        _primary_anchor_col = None
        _secondary_anchor_col = None
        _selected_fields = selected_fields if len(selected_fields) > 0 else None

        wb = pyxl.load_workbook(filepath)

        primary_repr = info["primary"]
        primary_sheetname = primary_repr["excel_sheetname"]

        if primary_sheetname in wb.sheetnames:  # collection was not merged
            primary_df = pd.read_excel(
                filepath, sheet_name=primary_sheetname, index_col=None, header=0
            )
            _primary = Dataset(
                primary_df,
                id_col_name=primary_repr["id_col_name"],
                date_col_name=primary_repr["date_col_name"],
                id2_col_name=primary_repr["id2_col_name"],
                name=primary_repr["name"],
            )

        else:  # collection was merged
            merged_dset_sheet_name = get_option("excel.sheet_name.merged_results")
            if merged_dset_sheet_name not in wb.sheetnames:
                raise ParserError(
                    f"Excel sheet '{primary_sheetname}' nor "
                    f"'{merged_dset_sheet_name}' not found in file: {filepath}"
                )

            ws = wb[sheet_name]
            if ws["A2"].value == get_option("excel.row_index_header"):
                _merged_dset = pd.read_excel(
                    filepath, sheet_name=merged_dset_sheet_name, index_col=0, header=[0, 1]
                )
            else:
                _merged_dset = pd.read_excel(
                    filepath, sheet_name=merged_dset_sheet_name, index_col=None, header=[0, 1]
                )

            primary_name = info["primary"]["name"]
            primary_df = _merged_dset.xs(primary_name, axis="columns", level=0)
            _primary = Dataset(
                primary_df,
                id_col_name=info["primary"]["id_col_name"],
                date_col_name=info["primary"]["date_col_name"],
                id2_col_name=info["primary"]["id2_col_name"],
                name=primary_name,
            )

            for sec in info["secondary"]:
                secondary_name = sec["name"]
                secondary_tags = sec["tags"]

                if MergeableAnchoredList.tag_merged in secondary_tags:
                    secondary_df = _merged_dset.xs(secondary_name, axis="columns", level=0)
                    secondary_dset = Dataset(
                        secondary_df,
                        id_col_name=sec["id_col_name"],
                        date_col_name=sec["date_col_name"],
                        id2_col_name=sec["id2_col_name"],
                        name=secondary_name,
                    )
                    secondary_dset.clear_tags()
                    secondary_dset.drop_sys_cols()
                    _secondary.append(secondary_dset)

        for sec in info["secondary"]:
            secondary_tags = sec["tags"]
            if MergeableAnchoredList.tag_merged not in secondary_tags:
                secondary_dset = Dataset.from_excel_sheet(filepath, sec)
                secondary_dset.clear_tags()
                secondary_dset.drop_sys_cols()
                _secondary.append(secondary_dset)

        _primary_anchor_col = info["primary_anchor_col"]
        _secondary_anchor_col = info["secondary_anchor_col"]

        instance = cls(
            primary=_primary,
            secondary=_secondary,
            primary_anchor_col=_primary_anchor_col,
            secondary_anchor_col=_secondary_anchor_col,
            selected_fields=_selected_fields,
        )

        return instance
