import itertools

import openpyxl as pyxl
import pandas as pd
import tablib as tl


YELLOW = "00FFFF00"


def autofit_column_width(ws):
    """Autoadjust the column widths of a Worksheet

    :param ws: :class:`openpyxl.worksheet.worksheet.Worksheet` to adjust
    """
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        length = min((length + 2) * 1.2, 65)
        ws.column_dimensions[column_cells[0].column_letter].width = length


def get_column_index(ws, col_header: str = None):
    """Get column index of the ``col_header``, returning ``-1`` if not found

    :param ws: :class:`openpyxl.worksheet.worksheet.Worksheet`
    :param col_header: column header to find
    """
    if col_header is None:
        return -1

    for cell in ws[1]:
        if cell.value == col_header:
            return cell.column
    for cell in ws[2]:
        if cell.value == col_header:
            return cell.column

    return -1


def get_sheet_names(filepath_or_buffer):
    book = pyxl.load_workbook(filepath_or_buffer, read_only=True, data_only=True)
    sheetnames = book.sheetnames
    book.close()
    return sheetnames


def highlight_row(ws, row: int, color: str = YELLOW):
    """Highlight row a certain color

    :param ws: :class:`openpyxl.worksheet.worksheet.Worksheet`
    :param row: row index to highlight
    :param color: color to highlight in RGB hexadecimal format (e.g. "00FFFF00"=yellow)
    """
    rows_iter = ws.iter_rows(min_col=1, min_row=row, max_col=ws.max_column, max_row=row)
    for row in rows_iter:
        for cell in row:
            cell.fill = pyxl.styles.PatternFill("solid", fgColor=color)


def is_row_empty(ws, row_index, delete_if_empty=False):
    """Determine if a row is empty.

    :param delete_if_empty: If True, and row is empty, row is deleted.
                            defaults to False
    """
    for row in ws.iter_rows(min_row=row_index, max_row=row_index):
        empty = not any((cell.value for cell in row))
        if empty:
            if delete_if_empty:
                ws.delete_rows(row_index, 1)
            return True
    return False


def iter_rows_with_column_value(ws, column: str, value):
    col_index = get_column_index(ws, column)

    if col_index > -1:
        # built-in constants should be compared with 'is'
        if value is None or isinstance(value, (bool)):
            predicate = lambda x: x is value
        else:
            predicate = lambda x: x == value

        rows_iter = ws.iter_rows(
            min_col=col_index, min_row=1, max_col=col_index, max_row=ws.max_row
        )

        for row in rows_iter:
            cell = row[0]
            if predicate(cell.value):
                yield cell.row


def to_df(ws, num_header: int = 1, num_idx: int = 0):
    """Converts an Excel worksheet to a :class:`pandas.DataFrame`.
    Better to use :func:`pandas.read_excel` as it takes care of
    more nuances.

    :param num_header: number of header rows
    :param num_idx: number of index columns
    """
    if num_header == 1 and num_idx == 0:
        data = ws.values
        cols = next(data)
        return pd.DataFrame(data, columns=cols)

    data = ws.values
    max_col_index = ws.max_column

    if num_header == 1:
        cols = next(data)[num_idx:]
    elif num_header > 1:
        col_tuples = []
        for x in range(num_header):
            col_tuples.append(next(data)[num_idx:])
        zipped = zip(*col_tuples)
        cols = pd.MultiIndex.from_tuples(list(zipped))
    else:
        # mimick pd.DataFrame constructor for when columns=None,
        # making room for Index/MultiIndex names
        cols = pd.RangeIndex(start=num_idx, stop=max_col_index)

    data = list(data)

    if num_idx == 1:
        idx = pd.Index([r[0] for r in data])
        if num_header == 0:
            idx.name = 0
    elif num_idx > 1:
        tuples = [r[:num_idx] for r in data]
        idx = pd.MultiIndex.from_tuples(tuples)
        if num_header == 0:
            idx.names = list(range(0, num_idx))
    else:
        idx = None

    data = (itertools.islice(r, num_idx, None) for r in data)

    df = pd.DataFrame(data, index=idx, columns=cols)

    return df


def to_tablib_dataset(wb, sheet_name=None, headers=True, skip_lines=0):
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]

    dset = tl.Dataset()
    dset.title = ws.title

    for i, row in enumerate(ws.rows):
        if i < skip_lines:
            continue
        row_vals = [c.value for c in row]
        if i == skip_lines and headers:
            dset.headers = row_vals
        else:
            dset.append(row_vals)

    return dset
