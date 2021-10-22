import openpyxl as pyxl

from . import listlike as lltools
from . import openpyxl as openpyxltools


def keep_ws(filepath, ws_to_keep):
    """Keep specified Excel worksheets (and discard the rest)."""
    wb = pyxl.load_workbook(filepath)
    ws_to_keep = lltools.maybe_make_list(ws_to_keep)
    ws_to_delete = lltools.diff(wb.sheetnames, ws_to_keep)
    for ws in ws_to_delete:
        del wb[ws]
    wb.save(filepath)


def move_ws_before_sheetname(filepath, ws, predicate=None):
    """Move Excel worksheet before the first worksheet name
    that starts with ``sheetname_pattern``.

    :param ws: Worksheet to move
    :param sheetname_pattern: ``ws`` is moved before the first worksheet name
                              that starts value of this parameter
    """
    if predicate is None:
        predicate = lambda ws: ws.title.startswith("_")

    wb = pyxl.load_workbook(filepath)
    insert_before_ws = next(ws for ws in wb.worksheets if predicate(ws))
    openpyxltools.wb_move_sheets(wb, ws, insert_before_ws.title)
    wb.save(filepath)
