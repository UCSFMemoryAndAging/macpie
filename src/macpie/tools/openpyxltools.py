import collections
import itertools
import re

import openpyxl as pyxl
import pandas as pd
import tablib as tl
from pandas.core.dtypes.common import is_re_compilable

from macpie.tools import lltools, strtools


YELLOW = "00FFFF00"


def autofit_column_width(ws):
    """Autoadjust the column widths of a Worksheet

    :param ws: :class:`openpyxl.worksheet.worksheet.Worksheet` to adjust
    """
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        length = min((length + 2) * 1.2, 65)
        ws.column_dimensions[column_cells[0].column_letter].width = length


def get_column_index(ws, col_header: str = None):
    """Get column index of the ``col_header``, returning ``-1`` if not found

    :param ws: :class:`openpyxl.worksheet.worksheet.Worksheet`
    :param col_header: column header to find
    """
    if col_header is None:
        return -1

    for cell in ws[1]:
        if cell.value == col_header:
            return cell.column
    for cell in ws[2]:
        if cell.value == col_header:
            return cell.column

    return -1


def get_sheet_names(filepath_or_buffer):
    """Get all sheet names from an Excel file."""
    book = pyxl.load_workbook(filepath_or_buffer)
    sheetnames = book.sheetnames
    return sheetnames


def highlight_row(ws, row: int, color: str = YELLOW):
    """Highlight row a certain color

    :param ws: :class:`openpyxl.worksheet.worksheet.Worksheet`
    :param row: row index to highlight
    :param color: color to highlight in RGB hexadecimal format (e.g. "00FFFF00"=yellow)
    """
    rows_iter = ws.iter_rows(min_col=1, min_row=row, max_col=ws.max_column, max_row=row)
    for row in rows_iter:
        for cell in row:
            cell.fill = pyxl.styles.PatternFill("solid", fgColor=color)


def is_row_empty(ws, row_index, delete_if_empty=False):
    """Determine if a row is empty.

    :param delete_if_empty: If True, and row is empty, row is deleted.
                            defaults to False
    """
    for row in ws.iter_rows(min_row=row_index, max_row=row_index):
        empty = not any((cell.value for cell in row))
        if empty:
            if delete_if_empty:
                ws.delete_rows(row_index, 1)
            return True
    return False


def iter_cells(
    ws, min_row=None, max_row=None, min_col=None, max_col=None, values_only=False, by="column"
):
    if by == "column":
        for col in ws.iter_cols(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=values_only,
        ):
            for cell in col:
                yield cell
    else:
        for col in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=values_only,
        ):
            for cell in col:
                yield cell


def iter_rows_with_column_value(ws, column: str, value):
    col_index = get_column_index(ws, column)

    if col_index > -1:
        # built-in constants should be compared with 'is'
        if value is None or isinstance(value, (bool)):
            predicate = lambda x: x is value
        else:
            predicate = lambda x: x == value

        rows_iter = ws.iter_rows(
            min_col=col_index, min_row=1, max_col=col_index, max_row=ws.max_row
        )

        for row in rows_iter:
            cell = row[0]
            if predicate(cell.value):
                yield cell.row


def iter_filepaths(filepaths, sheet_names=None):
    if sheet_names and lltools.is_list_like(sheet_names) and lltools.is_list_like(sheet_names[0]):
        if len(filepaths) != len(sheet_names):
            raise ValueError("filepaths and sheet_names have different lengths.")

    for idx, filepath in enumerate(filepaths):
        wb = pyxl.load_workbook(filepath)

        sheet_names_result = []
        if not sheet_names:
            sheet_names_result = wb.sheetnames
        elif isinstance(sheet_names, str):
            if sheet_names in wb.sheetnames:
                sheet_names_result = [sheet_names]
        elif lltools.is_list_like(sheet_names):
            if lltools.is_list_like(sheet_names[0]):
                sheet_names_for_file = lltools.maybe_make_list(sheet_names[idx])
                sheet_names_result = [
                    sheet_name
                    for sheet_name in sheet_names_for_file
                    if sheet_name in wb.sheetnames
                ]
            else:
                sheet_names_result = [
                    sheet_name for sheet_name in sheet_names if sheet_name in wb.sheetnames
                ]

        yield filepath, wb, sheet_names_result


def replace(ws, to_replace, value, ignorecase=False, regex=False, flags=0):
    """For a given Worksheet, replace values given in `to_replace` with `value`.

    Parameters
    ----------
    ws : :class:`openpyxl.worksheet.worksheet.Worksheet`
        Worksheet to replace values in
    to_replace : str or uncompiled regex
        String can be a character sequence or regular expression.
    value : str
        Replacement string. Will be coerced into same type as original cell value
    ignorecase : bool, default False
        Determines if replace is case sensitive:
    regex : bool, default False
        Determines if the passed-in pattern in ``to_replace`` is a regular expression.
    flags : int, default 0 (no flags)
        Regex module flags, e.g. re.IGNORECASE.
    """
    counter = collections.Counter()
    if regex:
        if not is_re_compilable(to_replace):
            raise TypeError("Could not compile 'to_replace' to regex.")
        if ignorecase:
            flags |= re.IGNORECASE

        for cell in iter_cells(ws):
            cell_value = str(cell.value)
            match = re.fullmatch(to_replace, cell_value, flags=flags)
            if match:
                counter[match.group(0)] += 1
                if value == "":
                    cell.value = value
                else:
                    orig_data_type = type(cell.value)
                    try:
                        cell.value = orig_data_type(value)
                    except TypeError:
                        raise TypeError(
                            f"Can't cast replacement value '{value}' to same "
                            f"type of original value '{orig_data_type}'."
                        )
    else:
        for cell in iter_cells(ws):
            cell_value = str(cell.value)
            if strtools.str_equals(cell_value, to_replace, case_sensitive=not ignorecase):
                counter[cell.value] += 1
                cell.value = value
    return counter


def file_to_dataframe(filepath, sheet_name=None, read_only=True):
    try:
        wb = pyxl.load_workbook(filepath, read_only=read_only, data_only=True)
        if sheet_name is None:
            ws = wb.active
        else:
            ws = wb[sheet_name]
        return worksheet_to_dataframe(ws)
    finally:
        if read_only:
            wb.close()


def worksheet_to_dataframe(ws, num_header: int = 1, num_idx: int = 0):
    """Converts an Excel worksheet to a :class:`pandas.DataFrame`.
    Better to use :func:`pandas.read_excel` as it has more options
    and is faster.

    Parameters
    ----------
    num_header : int
        Number of header rows
    num_idx : int
        Number of index columns
    """
    if num_header == 1 and num_idx == 0:
        data = ws.values
        cols = next(data)
        return pd.DataFrame(data, columns=cols)

    data = ws.values
    max_col_index = ws.max_column

    if num_header == 1:
        cols = next(data)[num_idx:]
    elif num_header > 1:
        col_tuples = []
        for x in range(num_header):
            col_tuples.append(next(data)[num_idx:])
        zipped = zip(*col_tuples)
        cols = pd.MultiIndex.from_tuples(list(zipped))
    else:
        # mimick pd.DataFrame constructor for when columns=None,
        # making room for Index/MultiIndex names
        cols = pd.RangeIndex(start=num_idx, stop=max_col_index)

    data = list(data)

    if num_idx == 1:
        idx = pd.Index([r[0] for r in data])
        if num_header == 0:
            idx.name = 0
    elif num_idx > 1:
        tuples = [r[:num_idx] for r in data]
        idx = pd.MultiIndex.from_tuples(tuples)
        if num_header == 0:
            idx.names = list(range(0, num_idx))
    else:
        idx = None

    data = (itertools.islice(r, num_idx, None) for r in data)

    df = pd.DataFrame(data, index=idx, columns=cols)

    return df


def to_tablib_dataset(wb, sheet_name=None, headers=True, skip_lines=0, tablib_class=tl.Dataset):
    """Return a Tablib Dataset from an Excel file.

    Source: Adapted from https://github.com/jazzband/tablib/blob/418cc691a005a055a66b8e4f189f67ab675cb121/src/tablib/formats/_xlsx.py#L85
    """
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]

    dset = tablib_class()
    dset.title = ws.title

    for i, row in enumerate(ws.rows):
        if i < skip_lines:
            continue
        row_vals = [c.value for c in row]
        if i == skip_lines and headers:
            dset.headers = row_vals
        else:
            dset.append(row_vals)

    return dset
