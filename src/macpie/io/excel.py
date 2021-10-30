"""
Utilities for conversion to Excel representation.
"""
import json
import re

import openpyxl as pyxl
import pandas as pd
import tablib as tl

import macpie as mp
from macpie._config import get_option
from macpie import lltools, openpyxltools, tablibtools


DATASETS_SHEET_NAME = "_mp_datasets"
COLLECTION_SHEET_NAME = "_mp_collection"

ENGINES = ["openpyxl"]

INVALID_TITLE_REGEX = re.compile(r"[\\*?:/\[\]]")


def safe_xlsx_sheet_title(s, replace="-"):
    return re.sub(INVALID_TITLE_REGEX, replace, s)[:31]


def read_excel(
    io,
    as_collection=False,
    sheet_name=0,
    header=0,
    names=None,
    index_col=None,
    usecols=None,
    squeeze=None,
    dtype=None,
    engine=None,
    converters=None,
    true_values=None,
    false_values=None,
    skiprows=None,
    nrows=None,
    na_values=None,
    keep_default_na=True,
    na_filter=True,
    verbose=False,
    parse_dates=False,
    date_parser=None,
    thousands=None,
    comment=None,
    skipfooter=0,
    convert_float=None,
    mangle_dupe_cols=True,
    storage_options=None,
):

    should_close = False
    if not isinstance(io, MACPieExcelFile):
        should_close = True
        io = MACPieExcelFile(io, storage_options=storage_options)
    elif engine and engine != io.engine:
        raise ValueError(
            "Engine should not be specified when passing "
            "an ExcelFile - ExcelFile already has the engine set"
        )

    try:
        if as_collection:
            data = io.parse_collection()
        else:
            data = io.parse(
                sheet_name=sheet_name,
                header=header,
                names=names,
                index_col=index_col,
                usecols=usecols,
                squeeze=squeeze,
                dtype=dtype,
                converters=converters,
                true_values=true_values,
                false_values=false_values,
                skiprows=skiprows,
                nrows=nrows,
                na_values=na_values,
                keep_default_na=keep_default_na,
                na_filter=na_filter,
                verbose=verbose,
                parse_dates=parse_dates,
                date_parser=date_parser,
                thousands=thousands,
                comment=comment,
                skipfooter=skipfooter,
                convert_float=convert_float,
                mangle_dupe_cols=mangle_dupe_cols,
            )
    finally:
        # make sure to close opened file handles
        if should_close:
            io.close()
    return data


class MACPieExcelFile(pd.io.excel._base.ExcelFile):
    def __init__(self, path_or_buffer, storage_options=None):
        self._engines["openpyxl"] = MACPieExcelReader

        super().__init__(path_or_buffer, engine="openpyxl", storage_options=storage_options)

        self._dataset_dicts = self.get_dataset_dicts()
        self._mi_dataset_dicts = self.get_mi_dataset_dicts()
        self._collection_dict = self.get_collection_dict()

    def __repr__(self):
        return (
            f"{self.__class__.__name__}("
            f"io={self._io!r}, "
            f"datasets={self.get_dataset_names()!r}, "
            f"collection_class={self.get_collection_class_name()!r})"
        )

    def get_dataset_dicts(self):
        if DATASETS_SHEET_NAME not in self.sys_sheet_names:
            return None
            # raise ValueError(f"Cannot read Excel file without a '{DATASETS_SHEET_NAME}' sheet")

        dataset_dicts = self._reader.parse_excel_dict(DATASETS_SHEET_NAME)
        return dataset_dicts

    def get_collection_dict(self):
        if COLLECTION_SHEET_NAME not in self.sys_sheet_names:
            return None

        collection_dict = self._reader.parse_excel_dict(COLLECTION_SHEET_NAME)
        return collection_dict

    def get_mi_dataset_dicts(self):
        if self._dataset_dicts is None:
            return None

        mi_datasets = {}
        for dset_sheetname, dset_excel_dict in self._dataset_dicts.items():
            if lltools.is_list_like(dset_excel_dict.get("id_col_name")):
                # indicates multiindex columns
                mi_datasets[dset_sheetname] = dset_excel_dict
        return mi_datasets

    def get_dataset_names(self):
        return list(self._dataset_dicts.keys())

    def get_collection_class_name(self):
        if self._collection_dict is None:
            return None
        return self._collection_dict["class_name"]

    def parse_multiindex_df(self, sheet_name):
        return self._reader.parse_multiindex_df(sheet_name)

    def parse_multiindex_dataset(self, sheet_name):
        return self._reader.parse_multiindex_dataset(sheet_name, self._dataset_dicts[sheet_name])

    def parse_simple_dataset(self, sheet_name):
        tlset = self._reader.parse_tablib_dataset(sheet_name)
        return tablibtools.SimpleDataset.from_tlset(tlset)

    def parse_dictlike_dataset(self, sheet_name):
        tlset = self._reader.parse_tablib_dataset(sheet_name)
        return tablibtools.DictLikeDataset.from_tlset(tlset)

    def parse_dataset_fields(self, sheet_name):
        sdset = self._reader.parse_simple_dataset(sheet_name)
        dataset_fields = mp.DatasetFields()
        sdset.df.apply(
            lambda x: dataset_fields.append_series(x, with_tags=True, tag_value="x"),
            axis="columns",
        )
        return dataset_fields

    def parse_collection(self):
        if not self._collection_dict:
            raise ValueError(
                f"Cannot parse as collection without '{self.collection_sheet_name}' sheet."
            )

        collection_class_name = self._collection_dict["class_name"]
        collection_class = getattr(mp.collections, collection_class_name)
        return collection_class.from_excel_dict(self, self._collection_dict)

    def parse(
        self,
        sheet_name=0,
        header=0,
        names=None,
        index_col=None,
        usecols=None,
        squeeze=None,
        converters=None,
        true_values=None,
        false_values=None,
        skiprows=None,
        nrows=None,
        na_values=None,
        parse_dates=False,
        date_parser=None,
        thousands=None,
        comment=None,
        skipfooter=0,
        convert_float=None,
        mangle_dupe_cols=True,
        **kwds,
    ):
        """
        Parse specified sheet(s) into a DataFrame.
        Equivalent to read_excel(ExcelFile, ...)  See the read_excel
        docstring for more info on accepted parameters.
        Returns
        -------
        DataFrame or dict of DataFrames
            DataFrame from the passed in Excel file.
        """
        from macpie import Dataset

        mi_datasets_to_parse = []
        index_to_sheetname = {}

        if sheet_name is None:
            sheet_name = list(self._dataset_dicts.keys())

        if isinstance(sheet_name, list):
            for asheetname in list(sheet_name):
                if isinstance(asheetname, int):
                    sheetname_from_index = self._reader.get_sheetname_by_index(asheetname)
                    index_to_sheetname[asheetname] = sheetname_from_index
                    if sheetname_from_index in self._mi_dataset_dicts:
                        sheet_name.remove(asheetname)
                        mi_datasets_to_parse.append(sheetname_from_index)
                else:
                    if asheetname in self._mi_dataset_dicts:
                        sheet_name.remove(asheetname)
                        mi_datasets_to_parse.append(asheetname)
        else:
            if isinstance(sheet_name, int):
                sheet = self._reader.get_sheet_by_index(sheet_name)
                index_to_sheetname[sheet_name] = sheet.title
                sheet_name = sheet.title

            if self._mi_dataset_dicts and sheet_name in self._mi_dataset_dicts:
                return self.parse_multiindex_dataset(sheet_name)

        sheetname_to_index = {v: k for k, v in index_to_sheetname.items()}

        dfs_ret_val = self._reader.parse(
            sheet_name=sheet_name,
            header=header,
            names=names,
            index_col=index_col,
            usecols=usecols,
            squeeze=squeeze,
            converters=converters,
            true_values=true_values,
            false_values=false_values,
            skiprows=skiprows,
            nrows=nrows,
            na_values=na_values,
            parse_dates=parse_dates,
            date_parser=date_parser,
            thousands=thousands,
            comment=comment,
            skipfooter=skipfooter,
            convert_float=convert_float,
            mangle_dupe_cols=mangle_dupe_cols,
            **kwds,
        )

        if type(dfs_ret_val) is dict:
            ret_val = {}

            for sheet_name in dfs_ret_val:
                df = dfs_ret_val[sheet_name]

                if isinstance(sheet_name, int):
                    sheet_name = index_to_sheetname[sheet_name]

                if sheet_name in self.sheet_names:
                    excel_dict = self._dataset_dicts.get(sheet_name)
                    dset = Dataset.from_excel_dict(excel_dict, df)

                    if sheet_name in sheetname_to_index:
                        ret_val[sheetname_to_index[sheet_name]] = dset
                    else:
                        ret_val[sheet_name] = dset

            for mi_dataset in mi_datasets_to_parse:
                dset = self._reader.parse_multiindex_dataset(
                    mi_dataset, self._dataset_dicts[mi_dataset]
                )

                if mi_dataset in sheetname_to_index:
                    ret_val[sheetname_to_index[mi_dataset]] = dset
                else:
                    ret_val[mi_dataset] = dset

            return ret_val
        else:
            excel_dict = self._dataset_dicts.get(sheet_name)
            if excel_dict is None:
                raise ValueError(
                    f"No dataset info for sheet '{sheet_name}' found "
                    f"in '{DATASETS_SHEET_NAME}' sheet."
                )

            return Dataset(
                data=dfs_ret_val,
                id_col_name=excel_dict.get("id_col_name"),
                date_col_name=excel_dict.get("date_col_name"),
                id2_col_name=excel_dict.get("id2_col_name"),
                name=excel_dict.get("name"),
                tags=excel_dict.get("tags"),
            )

    @property
    def sheet_names(self):
        return [
            sheet_name for sheet_name in self._reader.sheet_names if not sheet_name.startswith("_")
        ]

    @property
    def sys_sheet_names(self):
        return [
            sheet_name for sheet_name in self._reader.sheet_names if sheet_name.startswith("_")
        ]


class MACPieExcelReader(pd.io.excel._openpyxl.OpenpyxlReader):
    def get_sheetname_by_index(self, index):
        return self.get_sheet_by_index(index).title

    def load_workbook(self, filepath_or_buffer):
        # Closes an xlsx file in read-only mode
        # https://stackoverflow.com/questions/31416842/openpyxl-does-not-close-excel-workbook-in-read-only-mode
        import io

        if isinstance(filepath_or_buffer, io.BufferedReader):
            return super().load_workbook(filepath_or_buffer)

        with open(filepath_or_buffer, "rb") as f:
            in_mem_file = io.BytesIO(f.read())

        return pyxl.load_workbook(in_mem_file, read_only=True, data_only=True, keep_links=False)

    def parse_excel_dict(self, sheet_name, headers=True):
        ws = self.book.active if sheet_name is None else self.book[sheet_name]
        df = openpyxltools.to_df(ws)
        df = df.applymap(json.loads)
        dld = tablibtools.DictLikeDataset.from_df(df)
        return dld.to_dict()

    def parse_tablib_dataset(self, sheet_name, headers=True):
        ws = self.book.active if sheet_name is None else self.book[sheet_name]
        return openpyxltools.to_tablib_dataset(ws)

    def parse_simple_dataset(self, sheet_name, headers=True):
        tlset = self.parse_tablib_dataset(sheet_name, headers)
        return tablibtools.SimpleDataset.from_tlset(tlset)

    def parse_dictlike_dataset(self, sheet_name, headers=True):
        tlset = self.parse_tablib_dataset(sheet_name, headers)
        return tablibtools.DictLikeDataset.from_tlset(tlset)

    def parse_multiindex_df(self, sheet_name):
        ws = self.book[sheet_name]
        if ws["A2"].value == get_option("excel.row_index_header"):
            return self.parse(sheet_name=sheet_name, index_col=0, header=[0, 1])
        else:
            return self.parse(sheet_name=sheet_name, index_col=None, header=[0, 1])

    def parse_multiindex_dataset(self, sheet_name, excel_dict):
        from macpie import Dataset

        df = self.parse_multiindex_df(sheet_name)
        dset = Dataset.from_excel_dict(excel_dict, df)
        return dset


class MACPieExcelWriter(pd.io.excel._OpenpyxlWriter):
    def _get_sheet_name(self, sheet_name):
        if not sheet_name:
            sheet_name = get_option("excel.sheet_name.default")
        return super()._get_sheet_name(sheet_name)

    def write_excel_dict(self, excel_dict: dict):
        if excel_dict["class_name"] == "Dataset":
            sheet_name = DATASETS_SHEET_NAME
            excel_dict = {excel_dict["excel_sheetname"]: excel_dict}
        else:
            sheet_name = COLLECTION_SHEET_NAME

        dld = tablibtools.DictLikeDataset.from_dict(excel_dict)

        if sheet_name in self.book.sheetnames:
            ws = self.book[sheet_name]
        else:
            ws = self.book.create_sheet()
            ws.title = sheet_name
            ws.append(dld.headers)

        for row in dld.data:
            ws.append([json.dumps(cell) for cell in row])

    def write_tablib_dataset(self, tlset: tl.Dataset):
        ws = self.book.create_sheet()
        ws.title = (
            safe_xlsx_sheet_title(tlset.title, "-")
            if tlset.title
            else (get_option("excel.sheet_name.default"))
        )

        from tablib.formats._xlsx import XLSXFormat

        XLSXFormat.dset_sheet(tlset, ws)

    def write_simple_dataset(self, simple_dataset: tablibtools.SimpleDataset):
        self.write_tablib_dataset(simple_dataset.tlset)

    def handle_multiindex(self, sheet_name):
        ws = self.book[sheet_name]
        if openpyxltools.is_row_empty(ws, row_index=3, delete_if_empty=True):
            # Special case to handle pandas and openpyxl bugs when writing
            # dataframes with multiindex.
            # https://stackoverflow.com/questions/54682506/openpyxl-in-python-delete-rows-function-breaks-the-merged-cell
            # https://github.com/pandas-dev/pandas/issues/27772
            # Another openpyxl bug where if number of index cols > 1,
            # deleting rows doesn't work if adjacent cells in the index have been merged.
            # Since we are forced to keep the index column due to bug,
            # might as well give it an informative name
            ws["A2"].value = get_option("excel.row_index_header")

    def highlight_duplicates(self, sheet_name, column_name):
        ws = self.book[sheet_name]
        rows_to_highlight = openpyxltools.iter_rows_with_column_value(ws, column_name, True)
        for row in rows_to_highlight:
            openpyxltools.highlight_row(ws, row)

    def reorder_sheets(self):
        try:
            sheetnames = list(self.book.sheetnames)
            dset_sheet_index = sheetnames.index(DATASETS_SHEET_NAME) + 1
            sheetname_to_move_to = next(
                sheetname
                for sheetname in sheetnames[dset_sheet_index:]
                if sheetname.startswith("_mp")
            )
            lltools.move_item_to(sheetnames, DATASETS_SHEET_NAME, sheetname_to_move_to)
            self.book._sheets = [self.book[sheetname] for sheetname in sheetnames]
        except (ValueError, StopIteration) as e:
            pass

    def save(self):
        for ws in self.book.worksheets:
            if ws.title.startswith("_mp"):
                openpyxltools.autoadjust_column_widths(ws)

        self.reorder_sheets()

        super().save()


pd.io.excel.register_writer(MACPieExcelWriter)
