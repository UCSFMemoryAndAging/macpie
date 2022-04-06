import abc
import re

import openpyxl as pyxl
import pandas as pd
import tablib as tl

import macpie as mp
from macpie._config import get_option
from macpie import lltools, tablibtools


DATASETS_SHEET_NAME = "_mp_datasets"
COLLECTION_SHEET_NAME = "_mp_collection"

INVALID_TITLE_REGEX = re.compile(r"[\\*?:/\[\]]")


def safe_xlsx_sheet_title(s, replace="-"):
    return re.sub(INVALID_TITLE_REGEX, replace, s)[:31]


def read_excel(
    io,
    as_collection=False,
    sheet_name=0,
    header=0,
    names=None,
    index_col=None,
    usecols=None,
    squeeze=None,
    dtype=None,
    engine=None,
    converters=None,
    true_values=None,
    false_values=None,
    skiprows=None,
    nrows=None,
    na_values=None,
    keep_default_na=True,
    na_filter=True,
    verbose=False,
    parse_dates=False,
    date_parser=None,
    thousands=None,
    comment=None,
    skipfooter=0,
    convert_float=None,
    mangle_dupe_cols=True,
    storage_options=None,
):

    should_close = False
    if not isinstance(io, MACPieExcelFile):
        should_close = True
        io = MACPieExcelFile(io, storage_options=storage_options)
    elif engine and engine != io.engine:
        raise ValueError(
            "Engine should not be specified when passing "
            "an ExcelFile - ExcelFile already has the engine set"
        )

    try:
        if as_collection:
            data = io.parse_collection()
        else:
            data = io.parse(
                sheet_name=sheet_name,
                header=header,
                names=names,
                index_col=index_col,
                usecols=usecols,
                squeeze=squeeze,
                dtype=dtype,
                converters=converters,
                true_values=true_values,
                false_values=false_values,
                skiprows=skiprows,
                nrows=nrows,
                na_values=na_values,
                keep_default_na=keep_default_na,
                na_filter=na_filter,
                verbose=verbose,
                parse_dates=parse_dates,
                date_parser=date_parser,
                thousands=thousands,
                comment=comment,
                skipfooter=skipfooter,
                convert_float=convert_float,
                mangle_dupe_cols=mangle_dupe_cols,
            )
    finally:
        # make sure to close opened file handles
        if should_close:
            io.close()

    return data


class MACPieExcelFile(pd.io.excel._base.ExcelFile):
    def __init__(self, path_or_buffer, storage_options=None):
        from ._openpyxl import MACPieOpenpyxlReader

        self._engines["mp_openpyxl"] = MACPieOpenpyxlReader

        super().__init__(path_or_buffer, engine="mp_openpyxl", storage_options=storage_options)

        self._dataset_dicts = self.get_dataset_dicts()
        self._collection_dict = self.get_collection_dict()

    def __repr__(self):
        return (
            f"{self.__class__.__name__}("
            f"io={self._io!r}, "
            f"datasets={self.dataset_sheetnames!r}, "
            f"collection_class={self.collection_classname!r})"
        )

    @property
    def dataset_sheetnames(self):
        if self._dataset_dicts:
            return list(self._dataset_dicts.keys())
        return []

    @property
    def collection_classname(self):
        if self._collection_dict is None:
            return None
        return self._collection_dict["class_name"]

    def get_dataset_dicts(self):
        if DATASETS_SHEET_NAME not in self.sys_sheet_names:
            return {}
            # raise ValueError(f"Cannot read Excel file without a '{DATASETS_SHEET_NAME}' sheet")

        dataset_dicts = self._reader.parse_excel_dict_sheet(DATASETS_SHEET_NAME)

        for excel_dict in dataset_dicts.values():
            excel_dict["id_col_name"] = lltools.make_tuple_if_list_like(
                excel_dict.get("id_col_name")
            )
            excel_dict["date_col_name"] = lltools.make_tuple_if_list_like(
                excel_dict.get("date_col_name")
            )
            excel_dict["id2_col_name"] = lltools.make_tuple_if_list_like(
                excel_dict.get("id2_col_name")
            )

        return dataset_dicts

    def get_collection_dict(self):
        if COLLECTION_SHEET_NAME not in self.sys_sheet_names:
            return {}

        collection_dict = self._reader.parse_excel_dict_sheet(COLLECTION_SHEET_NAME)
        return collection_dict

    def parse_tablib_dataset(self, sheet_name=None, headers=True):
        return self._reader.parse_tablib_dataset(sheet_name=sheet_name, headers=headers)

    def parse_simple_dataset(self, sheet_name=None, headers=True):
        tlset = self.parse_tablib_dataset(sheet_name=sheet_name, headers=headers)
        return tablibtools.TablibDataset.from_tlset(tlset)

    def parse_dictlike_dataset(self, sheet_name):
        return self._reader.parse_dictlike_dataset(sheet_name)

    def parse_tablib_datasets(self, sheet_name=None, headers=True):
        ret_dict = False

        if isinstance(sheet_name, list):
            sheets = sheet_name
            ret_dict = True
        elif sheet_name is None:
            sheets = self.sheet_names
            ret_dict = True
        else:
            sheets = [sheet_name]

        output = {}

        for asheetname in sheets:
            if isinstance(asheetname, str):
                sheetname = asheetname
            else:
                sheetname = self._reader.get_sheetname_by_index(asheetname)

            tlset = self.parse_tablib_dataset(sheet_name=sheetname, headers=headers)
            output[asheetname] = tlset

        if ret_dict:
            return output
        else:
            return output[asheetname]

    def parse_simple_datasets(self, sheet_name=None, headers=True):
        tlset = self.parse_tablib_datasets(sheet_name=sheet_name, headers=headers)
        if type(tlset) is dict:
            return {k: tablibtools.TablibDataset.from_tlset(v) for (k, v) in tlset.items()}
        else:
            return tablibtools.TablibDataset.from_tlset(tlset)

    def parse_dataset_fields(self, sheet_name):
        sdset = self.parse_simple_dataset(sheet_name)
        dataset_fields = mp.DatasetFields()
        sdset.df.apply(
            lambda x: dataset_fields.append_series(x, with_tags=True, tag_value="x"),
            axis="columns",
        )
        return dataset_fields

    def parse_collection(self):
        if not self._collection_dict:
            raise ValueError(
                f"Cannot parse as collection without '{self.collection_sheet_name}' sheet."
            )

        collection_class_name = self._collection_dict["class_name"]
        collection_class = getattr(mp.collections, collection_class_name)
        return collection_class.from_excel_dict(self, self._collection_dict)

    def parse(
        self,
        sheet_name=0,
        header=0,
        names=None,
        index_col=None,
        usecols=None,
        squeeze=None,
        converters=None,
        true_values=None,
        false_values=None,
        skiprows=None,
        nrows=None,
        na_values=None,
        parse_dates=False,
        date_parser=None,
        thousands=None,
        comment=None,
        skipfooter=0,
        convert_float=None,
        mangle_dupe_cols=True,
        **kwds,
    ):
        """
        Parse specified sheet(s) into a DataFrame.
        Equivalent to read_excel(ExcelFile, ...)  See the read_excel
        docstring for more info on accepted parameters.
        Returns
        -------
        DataFrame or dict of DataFrames
            DataFrame from the passed in Excel file.
        """

        ret_dict = False

        if isinstance(sheet_name, list):
            sheets = sheet_name
            ret_dict = True
        elif sheet_name is None:
            sheets = self.sheet_names
            ret_dict = True
        else:
            sheets = [sheet_name]

        output = {}

        for asheetname in sheets:
            if isinstance(asheetname, str):
                sheetname = asheetname
            else:
                sheetname = self._reader.get_sheetname_by_index(asheetname)

            excel_dict = self._dataset_dicts.get(sheetname)
            if excel_dict is not None:
                read_excel_kwargs = excel_dict.get("read_excel_kwargs")
            else:
                read_excel_kwargs = {}

            df = self._reader.parse(
                sheet_name=sheetname,
                header=read_excel_kwargs.get("header", header),
                names=read_excel_kwargs.get("names", names),
                index_col=read_excel_kwargs.get("index_col", index_col),
                usecols=read_excel_kwargs.get("usecols", usecols),
                squeeze=read_excel_kwargs.get("squeeze", squeeze),
                converters=read_excel_kwargs.get("converters", converters),
                true_values=read_excel_kwargs.get("true_values", true_values),
                false_values=read_excel_kwargs.get("false_values", false_values),
                skiprows=read_excel_kwargs.get("skiprows", skiprows),
                nrows=read_excel_kwargs.get("nrows", nrows),
                na_values=read_excel_kwargs.get("na_values", na_values),
                parse_dates=read_excel_kwargs.get("parse_dates", parse_dates),
                date_parser=read_excel_kwargs.get("date_parser", date_parser),
                thousands=read_excel_kwargs.get("thousands", thousands),
                comment=read_excel_kwargs.get("comment", comment),
                skipfooter=read_excel_kwargs.get("skipfooter", skipfooter),
                convert_float=read_excel_kwargs.get("convert_float", convert_float),
                mangle_dupe_cols=read_excel_kwargs.get("mangle_dupe_cols", mangle_dupe_cols),
                **kwds,
            )

            if excel_dict is not None:
                output[asheetname] = mp.Dataset.from_excel_dict(excel_dict, df)
            else:
                output[asheetname] = df

        if ret_dict:
            return output
        else:
            return output[asheetname]

    @property
    def sheet_names(self):
        return [
            sheet_name for sheet_name in self._reader.sheet_names if not sheet_name.startswith("_")
        ]

    @property
    def sys_sheet_names(self):
        return [
            sheet_name for sheet_name in self._reader.sheet_names if sheet_name.startswith("_")
        ]


class MACPieExcelReader(pd.io.excel._base.BaseExcelReader):
    @abc.abstractmethod
    def get_sheetname_by_index(self, index):
        pass

    @abc.abstractmethod
    def parse_excel_dict_sheet(self, sheet_name):
        pass

    @abc.abstractmethod
    def parse_tablib_dataset(self, sheet_name, headers=True):
        pass

    def parse_simple_dataset(self, sheet_name):
        tlset = self.parse_tablib_dataset(sheet_name)
        return tablibtools.TablibDataset.from_tlset(tlset)

    def parse_dictlike_dataset(self, sheet_name):
        tlset = self.parse_tablib_dataset(sheet_name)
        return tablibtools.DictLikeDataset.from_tlset(tlset)

    def load_workbook(self, filepath_or_buffer):
        # Closes an xlsx file in read-only mode
        # https://stackoverflow.com/questions/31416842/openpyxl-does-not-close-excel-workbook-in-read-only-mode
        import io

        if isinstance(filepath_or_buffer, io.BufferedReader):
            return super().load_workbook(filepath_or_buffer)

        with open(filepath_or_buffer, "rb") as f:
            in_mem_file = io.BytesIO(f.read())

        return pyxl.load_workbook(in_mem_file, read_only=True, data_only=True, keep_links=False)


class MACPieExcelWriter(pd.io.excel._base.ExcelWriter):
    def __new__(cls, *args, **kwargs):
        # only switch class if generic(MACPieExcelWriter)
        if cls is MACPieExcelWriter:
            engine = kwargs.pop("engine", None)
            if engine is None:
                engine = get_option("excel.writer.engine")
            if engine not in ("mp_xlsxwriter", "mp_openpyxl"):
                raise ValueError(
                    f"Invalid engine: '{engine}''. Only 'mp_xlsxwriter' or 'mp_openpyxl' supported."
                )

            cls = pd.io.excel._util.get_writer(engine)

        return object.__new__(cls)

    @abc.abstractmethod
    def write_excel_dict(self, excel_dict: dict):
        pass

    @abc.abstractmethod
    def write_tablib_dataset(self, tlset: tl.Dataset, freeze_panes=True):
        pass

    @abc.abstractmethod
    def highlight_duplicates(self, sheet_name, column_name):
        pass

    @abc.abstractmethod
    def finalize_sheet_order(self):
        pass

    @abc.abstractmethod
    def sheet_names(self):
        pass

    def write_simple_dataset(self, simple_dataset: tablibtools.TablibDataset):
        self.write_tablib_dataset(simple_dataset.tlset)

    def finalized_sheet_order(self, sheetnames):
        try:
            dset_sheet_index = sheetnames.index(DATASETS_SHEET_NAME) + 1
            sheetname_to_move_to = next(
                sheetname
                for sheetname in sheetnames[dset_sheet_index:]
                if sheetname.startswith("_mp")
            )
            lltools.move_item_to(sheetnames, DATASETS_SHEET_NAME, sheetname_to_move_to)
        except (ValueError, StopIteration) as e:
            pass

        return sheetnames

    def _get_sheet_name(self, sheet_name):
        if not sheet_name:
            sheet_name = get_option("excel.sheet_name.default")
        return super()._get_sheet_name(sheet_name)
