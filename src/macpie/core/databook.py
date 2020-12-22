from pathlib import Path, PurePath
from typing import ClassVar

import openpyxl as pyxl
import pandas as pd

from .datasheet import Datasheet


class Databook:
    """
    A collection of Datasheet objects similar to an Excel workbook.
    """

    SHEETNAME_SHEETS : ClassVar[str] = '_sheets'

    def __init__(self):
        self._sheets = []

    def __getitem__(self, index):
        if isinstance(index, int):
            return self._sheets[index]
        else:
            raise ValueError("%s is not a valid index." % index)

    def __repr__(self) -> str:
        return (
            f'{self.__class__.__name__}('
            f'sheets={self._sheets!r})'
        )

    def add_book(self, another_book: "Databook"):
        """
        Append a Databook to this Databook
        """
        for ds in another_book:
            self.add_sheet(ds)

    def add_sheet(self, datasheet: Datasheet):
        """
        Add a Datasheet to this Databook.
        """
        for d in self._sheets:
            if d.name == datasheet.name:
                raise ValueError(f"Datasheet '{d.name}' already exists")

        self._sheets.append(datasheet)

    def add_metadata_sheet(self, datasheet: Datasheet):
        """
        Add a Datasheet representing metadata to this Databook.
        Converts all values to a JSON string so that they can be
        JSON loaded later if needed.
        """
        datasheet.df = datasheet.df.mac.json_dumps_contents()
        self._sheets.append(datasheet)

    @staticmethod
    def read_metadata_sheet(filepath, sheetname, index=None, parse_json=True, to_dict=True):
        """
        Reads in a Datasheet representing metadata and covert to a dict or DataFrame.

        :param filepath: Filepath of the file containing the metadata Datasheet
        :param sheetname: Name of the metadata Datasheet
        :param index: Set an index on the DataFrame
        :param parse_json: Whether to parse values as JSON strings
        :param to_dict: Whether to return result as a dict

        """
        sheet_data = pd.read_excel(filepath, sheet_name=sheetname, engine='openpyxl')
        if parse_json is True:
            sheet_data = sheet_data.mac.json_loads_contents()
        if index is not None:
            sheet_data = sheet_data.set_index(index)
        if to_dict is True:
            return sheet_data.to_dict('index')
        return sheet_data

    def to_excel(self, filepath: Path = None, mode='w'):
        """
        Write entire Databook to an Excel file.

        :param filepath: fielpath to write to
        :param mode: File mode to use (write ``'w'`` or append ``'a``')
        """

        filepath = self._validate_filepath(filepath)
        writer = pd.ExcelWriter(str(filepath), engine='openpyxl', mode=mode)

        log_sheets = []
        for ds in self._sheets:
            log_sheets.append(ds.to_dict())
            ds.to_excel(writer)

        Datasheet(
            self.SHEETNAME_SHEETS,
            pd.DataFrame(log_sheets).mac.json_dumps_contents()
        ).to_excel(writer)

        writer.save()
        self._format_excel(filepath)

    def _format_excel(self, filepath):
        wb = pyxl.load_workbook(filepath)
        sheet_log_dict = Databook.read_metadata_sheet(filepath, self.SHEETNAME_SHEETS, index='sheetname')
        for sheetname, sheetprops in sheet_log_dict.items():
            ws = wb[sheetname]
            if not sheetname.startswith('_'):
                if sheetprops['display_header'] is True:
                    # create excel auto filters if there is a header
                    if sheetprops['display_index'] is False:
                        filter_range = ('A'
                                        + str(sheetprops['num_header_rows'])
                                        + ":"
                                        + pyxl.utils.get_column_letter(ws.max_column)
                                        + str(ws.max_row))
                    elif sheetprops['display_index'] is True:
                        filter_range = (pyxl.utils.get_column_letter(sheetprops['num_index_cols'] + 1)
                                        + str(sheetprops['num_header_rows'])
                                        + ":"
                                        + pyxl.utils.get_column_letter(ws.max_column)
                                        + str(ws.max_row))
                        if sheetprops['num_header_rows'] > 1 and sheetprops['num_index_cols'] <= 1:
                            # special case to handle pandas and openpyxl bugs when writing dataframes with multiindex
                            # https://stackoverflow.com/questions/54682506/openpyxl-in-python-delete-rows-function-breaks-the-merged-cell
                            # https://github.com/pandas-dev/pandas/issues/27772
                            # another openpyxl bug where if num_index_cols > 1, deleting rows doesn't work if adjacent
                            # cells in the index have been merged
                            row_to_delete = sheetprops['num_header_rows'] + 1
                            ws.delete_rows(row_to_delete)
                    ws.auto_filter.ref = filter_range
        wb.save(filepath)

    def _validate_filepath(self, filepath):
        if filepath is None:
            raise ValueError("'filepath' cannot be None")
        if isinstance(filepath, PurePath):
            filepath = filepath.resolve()
        elif isinstance(filepath, str):
            filepath = Path(filepath).resolve()
        else:
            raise ValueError(
                f'For argument "filepath" expected type path or string, received '
                f"type {type(filepath).__name__}."
            )
        return filepath
