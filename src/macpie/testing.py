"""
Public testing utility functions.
"""

import openpyxl as pyxl
import pandas as pd

from macpie import datetimetools
from macpie.tools.openpyxl import iter_cells


def assert_dfs_equal(
    left: pd.DataFrame,
    right: pd.DataFrame,
    cols_ignore=set(),
    cols_ignore_pat="$^",
    output_dir=None,
):
    """For testing equality of :class:`pandas.DataFrame` objects

    Parameters
    ----------
    left : DataFrame
    right : DataFrame
    cols_ignore : list-like, optional
        Columns to ignore
    cols_ignore_pat : Regular expression. Default is ``$^``
        Column names that match will be ignored. Default pattern is ``$^``
        which matches nothing so no columns are ignored.
    output_dir : Path, optional
        Directory to write row difference results to
    """

    if left.mac.equals(right, cols_ignore, cols_ignore_pat):
        return True

    right = left.mac.assimilate(right)

    # check columns
    (left_only_cols, right_only_cols) = left.mac.diff_cols(
        right, cols_ignore=cols_ignore, cols_ignore_pat=cols_ignore_pat
    )

    if left_only_cols != set() or right_only_cols != set():
        assert False, f"\nleft_only_cols: {left_only_cols}\nright_only_cols: {right_only_cols}"

    # check rows
    pd.testing.assert_index_equal(left.index, right.index)

    row_diffs = left.mac.diff_rows(right, cols_ignore=cols_ignore, cols_ignore_pat=cols_ignore_pat)

    if row_diffs.mac.row_count() > 0:
        if output_dir is not None:
            row_diffs_filename = (
                "row_diffs_" + datetimetools.current_datetime_str(ms=True) + ".xlsx"
            )
            row_diffs.to_excel(output_dir / row_diffs_filename, index=False)

        assert False, f"\nrow_diffs: {row_diffs}"


def assert_excels_equal(filepath_1, filepath_2):
    """
    For testing equality of :class:`openpyxl.workbook.workbook.Workbook` objects

    Parameters
    ----------
    filepath_1 : Path
        Path of left Excel file to compare
    filepath_2 : Path
        Path of right Excel file to compare
    """
    wb1 = pyxl.load_workbook(filepath_1)
    wb2 = pyxl.load_workbook(filepath_2)

    # same sheets?
    assert set(wb1.sheetnames) == set(wb2.sheetnames)

    for sheet in wb1.sheetnames:
        assert_excel_worksheet_equal(wb1[sheet], wb2[sheet])


def assert_excel_worksheet_equal(left_ws, right_ws):
    """
    For testing equality of :class:`openpyxl.worksheet.worksheet.Worksheet` objects

    Parameters
    ----------
    left_ws : Worksheet
    right_ws : Worksheet
    """

    # same range of data?
    assert left_ws.max_column == right_ws.max_column
    assert left_ws.max_row == right_ws.max_row

    # same data in each cell?
    for left_cell in iter_cells(left_ws):
        right_cell = right_ws[left_cell.coordinate]
        assert left_cell.value == right_cell.value, (
            f"'{left_ws.title}'.{left_cell.coordinate} [{left_cell.value}] != "
            f"'{right_ws.title}'.{right_cell.coordinate} [{right_cell.value}]"
        )
