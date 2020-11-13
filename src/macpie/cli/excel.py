import functools

import numpy as np
import openpyxl as pyxl
import pandas as pd
from pathlib import Path

from macpie.io import create_output_dir
from macpie.io import get_row_by_col_val, ws_autoadjust_colwidth, ws_get_col, ws_highlight_row
from macpie.util import add_suffix, validate_bool_kwarg


duplicates_sheet_name_suffix = '(DUPS)'
fields_available_sheet_name = '_fields_available'
index_col_header = 'Original_Order'
link_id_col = 'link_id'
log_dataobjects_sheet_name = '_log_dataobjects'
log_operations_sheet_name = '_log_operations'
merged_results_sheet_name = 'MERGED_RESULTS'
primary_sheet_name_suffix = '_anchor'
secondary_sheet_name_suffix = '_linked'
sheet_name_chars_limit = 31


def write_standard(func):
    @functools.wraps(func)
    def write_standard_wrapper(*args, **kwargs):
        Q = args[0]
        if Q.executed is False:
            raise RuntimeError("should not write results file without first calling 'execute()'")
        output_dir = kwargs['output_dir'] if 'output_dir' in kwargs else None
        results_dir = create_output_dir(output_dir, "results")
        results_file = results_dir / (results_dir.stem + '.xlsx')
        writer = pd.ExcelWriter(results_file, engine='openpyxl')
        kwargs['writer'] = writer

        func(*args, **kwargs)

        log_dataobjects = pd.DataFrame(Q.log_dataobjects)
        log_operations = pd.DataFrame(Q.log_operations)
        log_dataobjects.to_excel(excel_writer=writer, sheet_name=log_dataobjects_sheet_name, index=False)
        log_operations.to_excel(excel_writer=writer, sheet_name=log_operations_sheet_name, index=False)
        writer.save()
        return results_file
    return write_standard_wrapper


@write_standard
def write_keepone_results(Q, output_dir: Path = None, **kwargs):
    writer = kwargs['writer']
    nodes_with_operations = Q.get_all_node_data('operation')
    for node in nodes_with_operations:
        result = node['operation_result']
        result.to_excel(excel_writer=writer, sheet_name=node['name'], index=False)

    edges_with_operation_results = Q.get_all_edge_data('operation_result')
    for edge in edges_with_operation_results:
        sheet_name = edge['name']
        if edge['duplicates']:
            sheet_name = add_suffix(sheet_name, "(DUPS)", sheet_name_chars_limit)
        edge['operation_result'].to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)


@write_standard
def write_link_results(Q, output_dir: Path = None, merge: bool = True, **kwargs):
    merge = validate_bool_kwarg(merge, "merge")
    writer = kwargs['writer']

    fields_list = []
    fields_list_dups = []

    if merge is False:
        nodes_with_operations = Q.get_all_node_data('operation')
        for node in nodes_with_operations:
            sheet_name = node['name']
            result = node['operation_result']
            fields_list.extend([(sheet_name, col) for col in result.columns])
            result.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)

        edges_with_operation_results = Q.get_all_edge_data('operation_result')
        for edge in edges_with_operation_results:
            sheet_name = edge['name']
            if edge['duplicates']:
                sheet_name = add_suffix(sheet_name, duplicates_sheet_name_suffix, sheet_name_chars_limit)
                fields_list_dups.extend([(sheet_name, col) for col in edge['operation_result'].columns])
            else:
                fields_list.extend([(sheet_name, col) for col in edge['operation_result'].columns])
            edge['operation_result'].to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
    else:
        primary_node = Q.get_root_node()
        primary_result = Q.get_node(primary_node, 'operation_result')
        primary_result.columns = pd.MultiIndex.from_product([[primary_node.name], primary_result.columns])

        final_result = primary_result

        for left_node, right_node, operation_result in Q.g.edges.data('operation_result'):
            if operation_result is not None:
                if Q.g.edges[left_node, right_node]['duplicates']:
                    # put secondary results that have duplicates as separate worksheets after the results worksheet
                    sheet_name = Q.g.edges[left_node, right_node]['name']
                    sheet_name = add_suffix(sheet_name, duplicates_sheet_name_suffix, sheet_name_chars_limit)
                    operation_result.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
                    fields_list_dups.extend([(sheet_name, col) for col in operation_result.columns])
                else:
                    # merge all secondary results that do not have any duplicates
                    final_result = final_result.mac.merge(
                        operation_result,
                        left_on=[
                            (primary_node.name, primary_node.id2_col),
                            (primary_node.name, primary_node.date_col),
                            (primary_node.name, primary_node.id_col)
                        ],
                        # mac.date_proximity function adds the '_link' suffix
                        right_on=[col + '_link' for col in [primary_node.id2_col,
                                                            primary_node.date_col,
                                                            primary_node.id_col]],
                        add_indexes=(None, right_node.name)
                    )

        fields_list.extend(list(final_result.columns))

        final_result.index = np.arange(1, len(final_result) + 1)
        final_result.to_excel(excel_writer=writer, sheet_name=merged_results_sheet_name, index=True)

    fields_list_df = pd.DataFrame(data=fields_list, columns=['Worksheet', 'Column'])
    fields_list_df.to_excel(excel_writer=writer, sheet_name=fields_available_sheet_name, index=False)


def format_results_basic(f):
    filename = str(f)
    wb = pyxl.load_workbook(filename)

    for ws in wb.worksheets:
        if ws.title.startswith('_log_'):
            ws_autoadjust_colwidth(ws)
        else:
            ws_highlight_dupes(ws, '_duplicates')

    wb.save(filename)


def format_link_results_with_merge(f):
    filename = str(f)
    wb = pyxl.load_workbook(filename)

    for ws in wb.worksheets:
        if ws.title.endswith(duplicates_sheet_name_suffix):
            ws_highlight_dupes(ws, '_duplicates')
        elif ws.title == merged_results_sheet_name:
            ws_index = wb.index(ws)
            wb.move_sheet(ws, -ws_index)
            format_multiindex(ws)
        elif ws.title.startswith('_log_'):
            ws_autoadjust_colwidth(ws)

    wb.save(filename)


def format_multiindex(ws):
    # get row index where column A has value of 1 (index of the dataframe)
    row_index = get_row_by_col_val(ws, 0, 1)
    row_index = row_index - 1
    ws.delete_rows(row_index)

    # forced to keep the index column due to bug, so might as well give it a good name
    ws['A2'].value = index_col_header
    # https://stackoverflow.com/questions/54682506/openpyxl-in-python-delete-rows-function-breaks-the-merged-cell
    # ws.delete_cols(1,1)

    data_range = "A2:" + pyxl.utils.get_column_letter(ws.max_column) + str(ws.max_row)

    ws.auto_filter.ref = data_range


def read_multiindex(f):
    filename = str(f)
    wb = pyxl.load_workbook(filename)
    ws = wb.active

    if ws['A2'].value == index_col_header:
        return pd.read_excel(filename, index_col=0, header=[0, 1], engine='openpyxl')
    else:
        return pd.read_excel(filename, index_col=None, header=[0, 1], engine='openpyxl')


def ws_highlight_dupes(ws, dupes_col):
    dupes_col = ws_get_col(ws, dupes_col)
    if dupes_col > -1:
        rows_iter = ws.iter_rows(min_col=dupes_col, min_row=1, max_col=dupes_col, max_row=ws.max_row)
        for row in rows_iter:
            cell = row[0]
            if cell.value is True:
                ws_highlight_row(ws, cell.row)
