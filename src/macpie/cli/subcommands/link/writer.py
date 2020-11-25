
import json
from pathlib import Path
from typing import ClassVar

import pandas as pd
import numpy as np
import openpyxl as pyxl

from macpie import util

from ...core import CliBaseQueryResults, SingleSheet


class CliLinkResults(CliBaseQueryResults):
    SHEETNAME_FIELDS_AVAILABLE : ClassVar[str] = '_fields_available'
    SHEETNAME_MERGE_INFO : ClassVar[str] = '_merge_info'
    SHEETNAME_MERGED_RESULTS : ClassVar[str] = 'MERGED_RESULTS'
    SHEETNAME_SUFFIX_DUPLICATES : ClassVar[str] = '(DUPS)'
    SHEETNAME_SUFFIX_PRIMARY : ClassVar[str] = '_anchor'
    SHEETNAME_SUFFIX_SECONDARY : ClassVar[str] = '_linked'

    COL_HEADER_DUPLICATES: ClassVar[str] = '_duplicates'
    COL_HEADER_FIELDS_SELECTED : ClassVar[str] = 'Merge?'
    COL_HEADER_ROW_INDEX : ClassVar[str] = 'Original_Order'
    COL_HEADER_LINK_ID : ClassVar[str] = 'link_id'

    def __init__(self, cli_ctx, output_dir: Path = None):
        super().__init__(cli_ctx, output_dir)
        self.merge = util.validators.validate_bool_kwarg(cli_ctx.obj['options']['merge_results'], 'merge')
        self.fields_list = []
        self.fields_list_dups = []

    def pre_write(self, Q):
        primary_node = Q.get_root_node()
        primary_result = Q.get_node(primary_node, 'operation_result')
        self.fields_list = [(primary_node.name,
                             col,
                             'x' if col in (primary_node.id_col,
                                            primary_node.date_col,
                                            primary_node.id2_col
                                            ) else None
                             ) for col in primary_result.columns
                            ]

        if self.merge is False:
            primary_sheetname = util.string.add_suffix(primary_node.name,
                                                       self.SHEETNAME_SUFFIX_PRIMARY,
                                                       self.SHEETNAME_CHARS_LIMIT)
            self.ws.append(SingleSheet(primary_sheetname, primary_result, None))

            edges_with_operation_results = Q.get_all_edge_data('operation_result')
            for edge in edges_with_operation_results:
                sheetname = util.string.add_suffix(edge['name'],
                                                   self.SHEETNAME_SUFFIX_SECONDARY,
                                                   self.SHEETNAME_CHARS_LIMIT)
                if edge['duplicates']:
                    sheetname = util.string.add_suffix(sheetname,
                                                       self.SHEETNAME_SUFFIX_DUPLICATES,
                                                       self.SHEETNAME_CHARS_LIMIT)
                self.fields_list.extend([(edge['name'], col, None) for col in edge['operation_result'].columns])
                self.ws.append(SingleSheet(sheetname, edge['operation_result'], None))
        else:
            primary_result.columns = pd.MultiIndex.from_product([[primary_node.name], primary_result.columns])
            final_result = primary_result

            self.fields_list_dups = []
            for left_node, right_node, operation_result in Q.g.edges.data('operation_result'):
                if operation_result is not None:
                    if Q.g.edges[left_node, right_node]['duplicates']:
                        # put secondary results that have duplicates as separate worksheets after the results worksheet
                        sheetname = util.string.add_suffix(Q.g.edges[left_node, right_node]['name'],
                                                           self.SHEETNAME_SUFFIX_SECONDARY,
                                                           self.SHEETNAME_CHARS_LIMIT)
                        sheetname = util.string.add_suffix(sheetname,
                                                           self.SHEETNAME_SUFFIX_DUPLICATES,
                                                           self.SHEETNAME_CHARS_LIMIT)
                        self.ws.append(SingleSheet(sheetname, operation_result, None))
                        self.fields_list_dups.extend([(Q.g.edges[left_node, right_node]['name'], col, None)
                                                      for col in operation_result.columns])
                    else:
                        self.fields_list.extend([(Q.g.edges[left_node, right_node]['name'], col, None)
                                                 for col in operation_result.columns])

                        # merge all secondary results that do not have any duplicates
                        final_result = final_result.mac.merge(
                            operation_result,
                            left_on=[
                                (primary_node.name, primary_node.id2_col),
                                (primary_node.name, primary_node.date_col),
                                (primary_node.name, primary_node.id_col)
                            ],
                            # mac.date_proximity function adds the '_link' suffix
                            right_on=[col + '_link' for col in [primary_node.id2_col,
                                                                primary_node.date_col,
                                                                primary_node.id_col]],
                            add_indexes=(None, right_node.name)
                        )

            final_result.index = np.arange(1, len(final_result) + 1)
            self.ws.append(SingleSheet(self.SHEETNAME_MERGED_RESULTS, final_result, None))

    def write(self, Q):
        self.pre_write(Q)

        for ws in self.ws:
            if ws.sheetname == self.SHEETNAME_MERGED_RESULTS:
                ws.df.to_excel(excel_writer=self.writer, sheet_name=ws[0], index=True)
            else:
                ws.df.to_excel(excel_writer=self.writer, sheet_name=ws[0], index=False)

        self.post_write(Q)
        self.writer.save()

        if self.merge:
            self.format_file()
        else:
            super().format_file()

    def post_write(self, Q):
        fields_list = self.fields_list + self.fields_list_dups
        fields_list_cols = ['DataObject', 'Field', self.COL_HEADER_FIELDS_SELECTED]
        fields_list_df = pd.DataFrame(data=fields_list, columns=fields_list_cols)
        fields_list_df.to_excel(excel_writer=self.writer, sheet_name=self.SHEETNAME_FIELDS_AVAILABLE, index=False)

        primary_node = Q.get_root_node()
        secondary_do_names = [o['right_operand'] for o in Q.log_operations if o['operation'] == 'date_proximity']
        merge_info_data = [
            ('primary', primary_node.to_json()),
            ('secondary', json.dumps(secondary_do_names)),
            ('merged', json.dumps(self.merge))
        ]
        merge_info_df = pd.DataFrame(data=merge_info_data, columns=['param_name', 'param_value'])
        merge_info_df.to_excel(excel_writer=self.writer, sheet_name=self.SHEETNAME_MERGE_INFO, index=False)
        super().post_write(Q)

    def format_file(self):
        filename = str(self.results_file)
        wb = pyxl.load_workbook(filename)

        for ws in wb.worksheets:
            if ws.title.endswith(self.SHEETNAME_SUFFIX_DUPLICATES):
                util.pyxl.ws_highlight_rows_with_col(ws, CliLinkResults.COL_HEADER_DUPLICATES)
            elif ws.title == self.SHEETNAME_MERGED_RESULTS:
                self.format_merged_results(ws)
            elif ws.title.startswith('_'):
                util.pyxl.ws_autoadjust_colwidth(ws)

        wb.save(filename)

    @staticmethod
    def format_merged_results(ws):
        # get row index where column A has value of 1 (index of the dataframe)
        row_index = util.pyxl.get_row_by_col_val(ws, 0, 1)
        row_index = row_index - 1
        ws.delete_rows(row_index)

        # forced to keep the index column due to bug, so might as well give it a good name
        ws['A2'].value = CliLinkResults.COL_HEADER_ROW_INDEX
        # https://stackoverflow.com/questions/54682506/openpyxl-in-python-delete-rows-function-breaks-the-merged-cell
        # ws.delete_cols(1,1)

        data_range = "A2:" + pyxl.utils.get_column_letter(ws.max_column) + str(ws.max_row)

        ws.auto_filter.ref = data_range

    @staticmethod
    def read_merged_results(f, sheetname: str = SHEETNAME_MERGED_RESULTS):
        filename = str(f)
        wb = pyxl.load_workbook(filename)
        ws = wb[sheetname]
        if ws['A2'].value == CliLinkResults.COL_HEADER_ROW_INDEX:
            return pd.read_excel(filename, index_col=0, header=[0, 1], engine='openpyxl')
        else:
            return pd.read_excel(filename, index_col=None, header=[0, 1], engine='openpyxl')
