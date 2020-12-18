from collections import namedtuple
from pathlib import Path
from shutil import copy
from typing import ClassVar

import click
import openpyxl as pyxl
import numpy as np
import pandas as pd

from macpie import io, util
from macpie.core import Databook, Datasheet, LavaDataObject

from macpie.cli.base import ClickPath, CmdParams
from macpie.cli.common import allowed_file, format_basic
from macpie.cli.exceptions import MergeParserError
from macpie.cli.subcommands import link


@click.command()
@click.option('--keep-original/--no-keep-original',
              default=True)
@click.argument('primary',
                nargs=1,
                type=ClickPath(exists=True, file_okay=True, dir_okay=False))
@click.pass_context
def merge(ctx, keep_original, primary):
    invoker = ctx.obj
    invoker.command_name = ctx.info_name
    invoker.add_opt('keep_original', keep_original)
    invoker.add_arg('primary', primary)

    params = MergeCmdParams(invoker.get_opt('verbose'),
                            invoker.get_opt('id_col'),
                            invoker.get_opt('date_col'),
                            invoker.get_opt('id2_col'),
                            invoker.get_opt('keep_original'),
                            invoker.get_arg('primary'))

    merge_parser = MergeParser(params.opts, params.args)

    databook = merge_parser.get_results()
    databook.add_metadata_sheet(invoker.get_command_info())
    databook.add_metadata_sheet(invoker.get_system_info())

    # if we need to keep some worksheets, best to copy the original file
    # and delete the rest since no easy way to currently copy worksheets
    # across workbooks
    if len(merge_parser.secondary_to_keep) > 0:
        path_to_link_results_copy = copy(merge_parser.link_results, invoker.results_dir)
        Path(path_to_link_results_copy).rename(invoker.results_file)

        wb = pyxl.load_workbook(invoker.results_file)
        ws_to_delete = util.list.diff(wb.sheetnames, merge_parser.secondary_to_keep)
        for ws in ws_to_delete:
            del wb[ws]

        wb.save(invoker.results_file)
        databook.to_excel(invoker.results_file, mode='a')
        format_order(invoker.results_file, merge_parser.secondary_to_keep)
    else:
        databook.to_excel(invoker.results_file)

    format_basic(invoker.results_file)
    link.format_dups(invoker.results_file)
    link.format_merged_results(invoker.results_file)


def format_order(filepath, secondary_to_keep):
    # move sheets into correct order
    wb = pyxl.load_workbook(filepath)
    insert_before_ws = next(x for x in wb.worksheets if x.title.startswith('_'))
    util.pyxl.move_sheets(wb, secondary_to_keep, insert_before_ws.title)
    wb.save(filepath)


class MergeCmdParams(CmdParams):

    def __init__(self,
                 verbose,
                 id_col,
                 date_col,
                 id2_col,
                 keep_original,
                 primary) -> None:

        _primary = io.validate_filepath(primary, allowed_file)

        self._opts = {
            'verbose': verbose,
            'id_col': id_col,
            'date_col': date_col,
            'id2_col': id2_col,
            'keep_original': util.validators.validate_bool_kwarg(keep_original, 'keep_original'),
        }

        self._args = {
            'primary': _primary
        }

    @property
    def opts(self):
        return self._opts

    @property
    def args(self):
        return self._args


class MergeParser:

    LOG_ENTRY = namedtuple('CliMergeLogEntry', ['sheetname',
                                                'do_name',
                                                'do_type',
                                                'do_description',
                                                'rows',
                                                'cols'])

    SHEETNAME_LOG_MERGE : ClassVar[str] = '_log_merge'

    def __init__(self, opts, args):
        self.opts = opts
        self.args = args

        self.primary = None  # DataObject
        self.secondary_to_merge = []  # list of DataObjects
        self.secondary_with_dups = Databook()
        self.secondary_to_keep = []  # list of sheet names
        self.log_merge = []

        self.validate_link_results(args['primary'])
        self.parse_primary()
        self.parse_secondary()

    def validate_link_results(self, link_results):
        self.link_results = link_results
        self.link_results_filename = str(self.link_results)
        self.link_results_wb = pyxl.load_workbook(self.link_results_filename)

        fields_available = Databook.read_metadata_sheet(
            self.link_results_filename,
            link.SHEETNAME_FIELDS_AVAILABLE,
            parse_json=False,
            to_dict=False
        )

        merge_info = Databook.read_metadata_sheet(
            self.link_results_filename,
            link.SHEETNAME_MERGE_INFO,
            index='param_name'
        )

        # get selected fields
        self.fields_selected = fields_available.loc[
            fields_available[link.COL_HEADER_FIELDS_SELECTED].isin(['x', 'X'])
        ]
        if self.fields_selected.mac.num_rows() < 1:
            raise MergeParserError("No fields selected to merge.")

        self.primary_do_info = merge_info['primary']['param_value']
        self.secondary_do_names = merge_info['secondary']['param_value']
        self.merged = merge_info['merged']['param_value']

        self.link_id_col = self.primary_do_info['id_col'] + '_link'
        self.link_date_col = self.primary_do_info['date_col'] + '_link'
        self.link_id2_col = self.primary_do_info['id2_col'] + '_link'

    def parse_primary(self):
        if self.opts['verbose']:
            click.echo("Parsing primary data object...")

        # are some results already merged?
        if self.merged is True:
            # get the worksheet with the merged results
            primary_do_sheetname = link.SHEETNAME_MERGED_RESULTS
            ws = self.link_results_wb[primary_do_sheetname]
            if ws['A2'].value == link.COL_HEADER_ROW_INDEX:
                primary_df = pd.read_excel(self.link_results_filename,
                                           sheet_name=primary_do_sheetname,
                                           index_col=0,
                                           header=[0, 1],
                                           engine='openpyxl')
            else:
                primary_df = pd.read_excel(self.link_results_filename,
                                           sheet_name=primary_do_sheetname,
                                           index_col=None,
                                           header=[0, 1],
                                           engine='openpyxl')
        else:
            # get the primary data worksheet and add primary do name as super index
            primary_do_sheetname = self.primary_do_info['name'] + link.SHEETNAME_SUFFIX_PRIMARY
            primary_df = pd.read_excel(self.link_results_filename, sheet_name=primary_do_sheetname, engine='openpyxl')
            primary_df.columns = pd.MultiIndex.from_product([[self.primary_do_info['name']], primary_df.columns])

        self.primary = LavaDataObject(
            name=self.primary_do_info['name'],
            df=primary_df,
            id_col=(self.primary_do_info['name'], self.primary_do_info['id_col']),
            date_col=(self.primary_do_info['name'], self.primary_do_info['date_col']),
            id2_col=(self.primary_do_info['name'], self.primary_do_info['id2_col']),
            display_name=primary_do_sheetname
        )

        self.log_merge.append(
            MergeParser.LOG_ENTRY(
                sheetname=self.primary.display_name,
                do_name=self.primary.name,
                do_type='primary',
                do_description={
                    'id_col': self.primary.id_col,
                    'date_col': self.primary.date_col,
                    'id2_col': self.primary.id2_col,
                },
                rows=self.primary.df.mac.num_rows(),
                cols=self.primary.df.mac.num_cols()
            )
        )

    def parse_secondary(self):
        if self.opts['verbose']:
            click.echo("Parsing secondary data objects...")

        # get selected data objects
        selected_do_names = list(self.fields_selected['DataObject'].unique())

        # get only secondary data object needed for merge
        if self.primary_do_info['name'] in selected_do_names:
            selected_do_names.remove(self.primary_do_info['name'])

        # items in secondary_do_names that are not in selected_do_names
        # keep as-is
        unselected_do_names = util.list.diff(self.secondary_do_names, selected_do_names)

        # quick check
        if not set(selected_do_names).issubset(set(self.secondary_do_names)):
            raise MergeParserError("DataObject selected but not available.")

        # check to see if any worksheets with duplicates have had them removed
        # if so, they can be merged; if not, leave them as is
        merge_description = "merged"  # for log entry purposes
        for do_name in selected_do_names:
            secondary_sheetname = do_name + link.SHEETNAME_SUFFIX_SECONDARY
            secondary_sheetname_dup = secondary_sheetname + link.SHEETNAME_SUFFIX_DUPLICATES
            if secondary_sheetname in self.link_results_wb.sheetnames:
                secondary_do_df = pd.read_excel(self.link_results_filename,
                                                sheet_name=secondary_sheetname,
                                                engine='openpyxl')
            elif secondary_sheetname_dup in self.link_results_wb.sheetnames:
                secondary_do_df = pd.read_excel(self.link_results_filename,
                                                sheet_name=secondary_sheetname_dup,
                                                engine='openpyxl')
                # recheck to see if all dups got removed. if so, can merge
                secondary_do_df[link.COL_HEADER_DUPLICATES] = False
                secondary_do_df[link.COL_HEADER_DUPLICATES
                                ] = secondary_do_df.duplicated(subset=self.link_id_col, keep=False)
                if secondary_do_df[link.COL_HEADER_DUPLICATES].any():
                    self.secondary_with_dups.add_sheet(Datasheet(secondary_sheetname_dup, secondary_do_df))
                    self.log_merge.append(
                        MergeParser.LOG_ENTRY(
                            sheetname=secondary_sheetname_dup,
                            do_name=do_name,
                            do_type='secondary',
                            do_description='no merge: has duplicates',
                            rows=secondary_do_df.mac.num_rows(),
                            cols=secondary_do_df.mac.num_cols()
                        )
                    )
                    continue
                else:
                    merge_description = 'merged: had duplicates before but all duplicates have been removed'
            else:
                # these should already be in the MERGED_RESULTS worksheet so no need to merge them
                self.log_merge.append(
                    MergeParser.LOG_ENTRY(
                        sheetname=secondary_sheetname,
                        do_name=do_name,
                        do_type='secondary',
                        do_description='no merge: already in MERGED_RESULTS worksheet. ignoring',
                        rows=None,
                        cols=None
                    )
                )
                continue

            has_link_table_cols = all([self.link_id_col in secondary_do_df,
                                       self.link_date_col in secondary_do_df,
                                       self.link_id2_col in secondary_do_df])
            if not has_link_table_cols:
                raise MergeParserError("Secondary data object doesn't have necessary link columns")

            secondary_do = LavaDataObject(
                name=do_name,
                df=secondary_do_df,
                id_col=self.link_id_col,
                date_col=self.link_date_col,
                id2_col=self.link_id2_col
            )
            self.secondary_to_merge.append(secondary_do)

            self.log_merge.append(
                MergeParser.LOG_ENTRY(
                    sheetname=secondary_sheetname,
                    do_name=do_name,
                    do_type='secondary',
                    do_description=merge_description,
                    rows=secondary_do_df.mac.num_rows(),
                    cols=secondary_do_df.mac.num_cols()
                )
            )

        for do_name in unselected_do_names:
            secondary_sheetname = do_name + link.SHEETNAME_SUFFIX_SECONDARY
            secondary_sheetname_dup = secondary_sheetname + link.SHEETNAME_SUFFIX_DUPLICATES
            if secondary_sheetname in self.link_results_wb.sheetnames:
                self.secondary_to_keep.append(secondary_sheetname)
            elif secondary_sheetname_dup in self.link_results_wb.sheetnames:
                self.secondary_to_keep.append(secondary_sheetname_dup)
            else:
                continue

            self.log_merge.append(
                MergeParser.LOG_ENTRY(
                    sheetname=self.secondary_to_keep[-1],
                    do_name=do_name,
                    do_type='secondary',
                    do_description='no merge: not selected for merge',
                    rows=None,
                    cols=None
                )
            )

    def get_results(self):
        db = Databook()

        final_result = self.primary.df

        for secondary in self.secondary_to_merge:
            if self.opts['verbose']:
                click.echo(f"Merging {secondary.name}")
            final_result = final_result.mac.merge(
                secondary.df,
                left_on=[self.primary.id2_col, self.primary.date_col, self.primary.id_col],
                right_on=[self.link_id2_col, self.link_date_col, self.link_id_col],
                add_indexes=(None, secondary.name)
            )

        fields_selected_records = self.fields_selected.to_dict('records')
        fields_selected = [(field['DataObject'], field['Field']) for field in fields_selected_records]

        # for new fields_available metadata sheet
        pre_filled_fields_available = [tup + ('x',) for tup in fields_selected]

        # always include primary link fields
        primary_link_fields = [
            (self.primary_do_info['name'], self.primary_do_info['id2_col']),
            (self.primary_do_info['name'], self.primary_do_info['date_col']),
            (self.primary_do_info['name'], self.primary_do_info['id_col'])
        ]
        fields_selected = primary_link_fields + fields_selected

        if self.opts['verbose']:
            click.echo("Removing unwanted fields...")
        final_result = final_result.filter(items=fields_selected, axis='columns')

        if self.opts['verbose']:
            click.echo("Writing merged results...")
        final_result.index = np.arange(1, len(final_result) + 1)
        db.add_sheet(Datasheet(link.SHEETNAME_MERGED_RESULTS, final_result, display_index=True))

        for secondary in self.secondary_with_dups:
            if self.opts['verbose']:
                click.echo("Writing results that still have duplicates...")
            db.add_sheet(secondary)

        # metadata sheets
        db.add_sheet(link.LinkQuery.create_fields_available_sheet(pre_filled_fields_available))
        db.add_metadata_sheet(self.create_merge_info_sheet())
        db.add_metadata_sheet(Datasheet(self.SHEETNAME_LOG_MERGE, pd.DataFrame(self.log_merge)))

        return db

    def create_merge_info_sheet(self):
        # create merge info sheet
        merge_info_data = [
            ('primary', self.primary_do_info),
            ('secondary', self.secondary_do_names),
            ('merged', True)
        ]
        merge_info_df = pd.DataFrame(data=merge_info_data, columns=['param_name', 'param_value'])
        return Datasheet(link.SHEETNAME_MERGE_INFO, merge_info_df)
