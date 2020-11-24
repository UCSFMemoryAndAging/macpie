from collections import namedtuple
import json
from pathlib import Path
from shutil import copy
from typing import ClassVar

import click
import openpyxl as pyxl
import pandas as pd

from macpie.classes import LavaDataObject
from macpie.io import move_sheets, ws_autoadjust_colwidth, ws_highlight_rows_with_col
from macpie.util import list_diff

from ..link.writer import CliLinkResults
from ...core import CliBaseResults, SingleSheet
from ...exceptions import CliLinkResultsParserError


CliMergeLogEntry = namedtuple('CliMergeLogEntry', ['sheetname',
                                                   'do_name',
                                                   'do_type',
                                                   'do_description',
                                                   'rows',
                                                   'cols'])


class CliMergeResults(CliBaseResults):

    SHEETNAME_LOG_MERGE : ClassVar[str] = '_log_merge'

    def __init__(self, cli_ctx, output_dir: Path = None):
        super().__init__(cli_ctx, output_dir)

        self.primary = None  # a SingleSheet
        self.secondary_to_merge = []  # list of SingleSheet
        self.secondary_with_dups = []  # lisit of SingleSheet
        self.secondary_to_keep = []  # list of sheet_names
        self.log_merge = []

        self._validate_link_results_file()
        self.parse_primary()
        self.parse_secondary()

    def _validate_link_results_file(self):
        if self.verbose:
            click.echo("Validating file...")

        self.link_results = self.cli_ctx.obj['args']['primary']
        self.link_results_filename = str(self.link_results)
        self.link_results_wb = pyxl.load_workbook(self.link_results_filename)

        fields_available = pd.read_excel(self.link_results_filename,
                                         sheet_name=CliLinkResults.SHEETNAME_FIELDS_AVAILABLE,
                                         engine='openpyxl')
        merge_info = pd.read_excel(self.link_results_filename,
                                   sheet_name=CliLinkResults.SHEETNAME_MERGE_INFO,
                                   engine='openpyxl')

        # get selected fields
        self.fields_selected = fields_available.loc[
            fields_available[CliLinkResults.COL_HEADER_FIELDS_SELECTED].isin(['x', 'X'])
        ]
        if self.fields_selected.mac.num_rows() < 1:
            raise CliLinkResultsParserError("No fields selected to merge.")

        self.primary_do_info = json.loads(
            merge_info.loc[merge_info['param_name'] == 'primary', 'param_value'].item()
        )
        self.secondary_do_names = json.loads(
            merge_info.loc[merge_info['param_name'] == 'secondary', 'param_value'].item()
        )
        self.merged = json.loads(
            merge_info.loc[merge_info['param_name'] == 'merged', 'param_value'].item()
        )

        self.link_id_col = self.primary_do_info['id_col'] + '_link'
        self.link_date_col = self.primary_do_info['date_col'] + '_link'
        self.link_id2_col = self.primary_do_info['id2_col'] + '_link'

    def parse_primary(self):
        if self.verbose:
            click.echo("Parsing primary data object...")

        # are some results already merged?
        if self.merged is True:
            # get the worksheet with the merged results
            primary_do_sheetname = CliLinkResults.SHEETNAME_MERGED_RESULTS
            ws = self.link_results_wb[primary_do_sheetname]
            if ws['A2'].value == CliLinkResults.COL_HEADER_ROW_INDEX:
                primary_df = pd.read_excel(self.link_results_filename,
                                           sheet_name=primary_do_sheetname,
                                           index_col=0,
                                           header=[0, 1],
                                           engine='openpyxl')
            else:
                primary_df = pd.read_excel(self.link_results_filename,
                                           sheet_name=primary_do_sheetname,
                                           index_col=None,
                                           header=[0, 1],
                                           engine='openpyxl')
        else:
            # get the primary data worksheet and add primary do name as super index
            primary_do_sheetname = self.primary_do_info['name'] + CliLinkResults.SHEETNAME_SUFFIX_PRIMARY
            primary_df = pd.read_excel(self.link_results_filename, sheet_name=primary_do_sheetname, engine='openpyxl')
            primary_df.columns = pd.MultiIndex.from_product([[self.primary_do_info['name']], primary_df.columns])

        primary_do = LavaDataObject(
            name=self.primary_do_info['name'],
            df=primary_df,
            id_col=(self.primary_do_info['name'], self.primary_do_info['id_col']),
            date_col=(self.primary_do_info['name'], self.primary_do_info['date_col']),
            id2_col=(self.primary_do_info['name'], self.primary_do_info['id2_col'])
        )
        self.primary = SingleSheet(primary_do_sheetname, primary_df, primary_do)

        self.log_merge.append(
            CliMergeLogEntry(
                sheetname=primary_do_sheetname,
                do_name=primary_do.name,
                do_type='primary',
                do_description=json.dumps({
                    'id_col': self.primary_do_info['id_col'],
                    'date_col': self.primary_do_info['date_col'],
                    'id2_col': self.primary_do_info['id2_col'],
                }),
                rows=primary_do.df.mac.num_rows(),
                cols=primary_do.df.mac.num_cols()
            )
        )

    def parse_secondary(self):
        if self.verbose:
            click.echo("Parsing secondary data objects...")

        # get selected data objects
        selected_do_names = list(self.fields_selected['DataObject'].unique())

        # get only secondary data object needed for merge
        if self.primary_do_info['name'] in selected_do_names:
            selected_do_names.remove(self.primary_do_info['name'])

        # items in secondary_do_names that are not in selected_do_names
        # keep as-is
        unselected_do_names = list_diff(self.secondary_do_names, selected_do_names)

        # quick check
        if not set(selected_do_names).issubset(set(self.secondary_do_names)):
            raise CliLinkResultsParserError("DataObject selected but not available.")

        # check to see if any worksheets with duplicates have had them removed
        # if so, they can be merged; if not, leave them as is
        for do_name in selected_do_names:
            secondary_sheetname = do_name + CliLinkResults.SHEETNAME_SUFFIX_SECONDARY
            secondary_sheetname_dup = secondary_sheetname + CliLinkResults.SHEETNAME_SUFFIX_DUPLICATES
            if secondary_sheetname in self.link_results_wb.sheetnames:
                secondary_do_df = pd.read_excel(self.link_results_filename,
                                                sheet_name=secondary_sheetname,
                                                engine='openpyxl')
            elif secondary_sheetname_dup in self.link_results_wb.sheetnames:
                secondary_do_df = pd.read_excel(self.link_results_filename,
                                                sheet_name=secondary_sheetname_dup,
                                                engine='openpyxl')
                # recheck to see if all dups got removed. if so, can merge
                secondary_do_df[CliLinkResults.COL_HEADER_DUPLICATES] = False
                secondary_do_df[CliLinkResults.COL_HEADER_DUPLICATES
                                ] = secondary_do_df.duplicated(subset=self.link_id_col, keep=False)
                if secondary_do_df[CliLinkResults.COL_HEADER_DUPLICATES].any():
                    self.secondary_with_dups.append(SingleSheet(secondary_sheetname_dup, secondary_do_df, None))
                    self.log_merge.append(
                        CliMergeLogEntry(
                            sheetname=secondary_sheetname_dup,
                            do_name=do_name,
                            do_type='secondary',
                            do_description='no merge: has duplicates',
                            rows=secondary_do_df.mac.num_rows(),
                            cols=secondary_do_df.mac.num_cols()
                        )
                    )
                    continue
                else:
                    self.log_merge.append(
                        CliMergeLogEntry(
                            sheetname=secondary_sheetname,
                            do_name=do_name,
                            do_type='secondary',
                            do_description='merged: had duplicates before but all duplicates have been removed',
                            rows=secondary_do_df.mac.num_rows(),
                            cols=secondary_do_df.mac.num_cols()
                        )
                    )
            else:
                # these should already be in the MERGED_RESULTS worksheet so no need to merge them
                self.log_merge.append(
                    CliMergeLogEntry(
                        sheetname=secondary_sheetname,
                        do_name=do_name,
                        do_type='secondary',
                        do_description='no merge: already in MERGED_RESULTS worksheet. ignoring',
                        rows=None,
                        cols=None
                    )
                )
                continue

            has_link_table_cols = all([self.link_id_col in secondary_do_df,
                                       self.link_date_col in secondary_do_df,
                                       self.link_id2_col in secondary_do_df])
            if not has_link_table_cols:
                raise CliLinkResultsParserError("Secondary data object doesn't have necessary link columns")

            secondary_do = LavaDataObject(
                name=do_name,
                df=secondary_do_df,
                id_col=self.link_id_col,
                date_col=self.link_date_col,
                id2_col=self.link_id2_col
            )
            self.secondary_to_merge.append(SingleSheet(secondary_sheetname, secondary_do_df, secondary_do))

            self.log_merge.append(
                CliMergeLogEntry(
                    sheetname=secondary_sheetname,
                    do_name=do_name,
                    do_type='secondary',
                    do_description='merged',
                    rows=secondary_do_df.mac.num_rows(),
                    cols=secondary_do_df.mac.num_cols()
                )
            )

        for do_name in unselected_do_names:
            secondary_sheetname = do_name + CliLinkResults.SHEETNAME_SUFFIX_SECONDARY
            secondary_sheetname_dup = secondary_sheetname + CliLinkResults.SHEETNAME_SUFFIX_DUPLICATES
            if secondary_sheetname in self.link_results_wb.sheetnames:
                self.secondary_to_keep.append(secondary_sheetname)
            elif secondary_sheetname_dup in self.link_results_wb.sheetnames:
                self.secondary_to_keep.append(secondary_sheetname_dup)
            else:
                continue

            self.log_merge.append(
                CliMergeLogEntry(
                    sheetname=self.secondary_to_keep[-1],
                    do_name=do_name,
                    do_type='secondary',
                    do_description='no merge: not selected for merge',
                    rows=None,
                    cols=None
                )
            )

    def check_ws_to_keep(self):
        # if we need to keep some worksheets, best to copy the original file
        # and delete the rest since no easy way to currently copy worksheets
        # across workbooks
        if len(self.secondary_to_keep) > 0:
            path_to_link_results_copy = copy(self.link_results, self.results_dir)
            Path(path_to_link_results_copy).rename(self.results_file)

            wb = pyxl.load_workbook(str(self.results_file))
            ws_to_delete = list_diff(wb.sheetnames, self.secondary_to_keep)
            for ws in ws_to_delete:
                del wb[ws]

            wb.save(str(self.results_file))
            self.writer = pd.ExcelWriter(self.results_file, mode='a', engine='openpyxl')

    def write(self):
        if self.verbose:
            click.echo("Performing merge...")
        self.check_ws_to_keep()

        final_result = self.primary.df

        for secondary in self.secondary_to_merge:
            if self.verbose:
                click.echo(f"Merging {secondary.do.name}")
            final_result = final_result.mac.merge(
                secondary.df,
                left_on=[self.primary.do.id2_col, self.primary.do.date_col, self.primary.do.id_col],
                right_on=[self.link_id2_col, self.link_date_col, self.link_id_col],
                add_indexes=(None, secondary.do.name)
            )

        fields_selected_records = self.fields_selected.to_dict('records')
        fields_selected = [(field['DataObject'], field['Field']) for field in fields_selected_records]
        if self.verbose:
            click.echo("Removing unwanted fields...")
        final_result = final_result.filter(items=fields_selected, axis='columns')

        if self.verbose:
            click.echo("Writing merged results...")
        final_result.to_excel(excel_writer=self.writer, sheet_name=CliLinkResults.SHEETNAME_MERGED_RESULTS, index=True)

        for secondary in self.secondary_with_dups:
            if self.verbose:
                click.echo("Writing results that still have duplicates...")
            secondary.df.to_excel(excel_writer=self.writer, sheet_name=secondary.sheetname, index=False)

        self.post_write()
        self.writer.save()
        self.format_file()

    def post_write(self):
        log_merge = pd.DataFrame(self.log_merge)
        log_merge.to_excel(excel_writer=self.writer, sheet_name=self.SHEETNAME_LOG_MERGE, index=False)
        super().post_write()

    def format_file(self):
        if self.verbose:
            click.echo("Begin formatting...")
        filename = str(self.results_file)
        wb = pyxl.load_workbook(filename)

        if len(self.secondary_to_keep) > 0:
            insert_before_ws = next(x for x in wb.worksheets if x.title.startswith('_'))
            move_sheets(wb, self.secondary_to_keep, insert_before_ws.title)

        for ws in wb.worksheets:
            if ws.title.endswith(CliLinkResults.SHEETNAME_SUFFIX_DUPLICATES):
                if self.verbose:
                    click.echo(f"Highlighting duplicates for {ws.title}...")
                ws_highlight_rows_with_col(ws, CliLinkResults.COL_HEADER_DUPLICATES)
            elif ws.title == CliLinkResults.SHEETNAME_MERGED_RESULTS:
                if self.verbose:
                    click.echo("Formatting merged results...")
                CliLinkResults.format_merged_results(ws)
            elif ws.title.startswith('_'):
                ws_autoadjust_colwidth(ws)

        wb.save(filename)
