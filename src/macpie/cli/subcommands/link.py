from functools import partial

import click
import numpy as np
import pandas as pd
import openpyxl as pyxl

from macpie import errors, io, pandas, util
from macpie.core import Databook, Datasheet, LavaDataObject, Query

from macpie.cli.base import ClickPath, CmdParams
from macpie.cli.common import allowed_file, format_basic


COL_HEADER_DUPLICATES = '_duplicates'
COL_HEADER_FIELDS_SELECTED = 'Merge?'
COL_HEADER_ROW_INDEX = 'Original_Order'
COL_HEADER_LINK_ID = 'link_id'

SHEETNAME_FIELDS_AVAILABLE = '_fields_available'
SHEETNAME_MERGE_INFO = '_merge_info'
SHEETNAME_MERGED_RESULTS = 'MERGED_RESULTS'
SHEETNAME_SUFFIX_DUPLICATES = '(DUPS)'
SHEETNAME_SUFFIX_PRIMARY = '_anchor'
SHEETNAME_SUFFIX_SECONDARY = '_linked'


@click.command()
@click.option('-k', '--primary-keep',
              default='all',
              type=click.Choice(['all', 'earliest', 'latest'], case_sensitive=False))
@click.option('-g', '--secondary-get',
              default='all',
              type=click.Choice(['all', 'closest'], case_sensitive=False))
@click.option('-t', '--secondary-days',
              default=90)
@click.option('-w', '--secondary-when',
              default='earlier_or_later',
              type=click.Choice(['earlier', 'later', 'earlier_or_later'], case_sensitive=False))
@click.option('-i', '--secondary-id-col',
              default=None,
              help="Secondary ID Column Header")
@click.option('-d', '--secondary-date-col',
              default=None,
              help="Secondary Date Column Header")
@click.option('-j', '--secondary-id2-col',
              default=None,
              help="Secondary ID2 Column Header")
@click.option('--merge-results/--no-merge-results',
              default=True)
@click.option('--keep-original/--no-keep-original',
              default=False)
@click.argument('primary',
                nargs=1,
                type=ClickPath(exists=True, file_okay=True, dir_okay=True))
@click.argument('secondary',
                nargs=-1,
                type=ClickPath(exists=True, file_okay=True, dir_okay=True))
@click.pass_context
def link(ctx,
         primary_keep,
         secondary_get,
         secondary_days,
         secondary_when,
         secondary_id_col,
         secondary_date_col,
         secondary_id2_col,
         merge_results,
         keep_original,
         primary,
         secondary):
    invoker = ctx.obj
    invoker.command_name = ctx.info_name
    invoker.add_opt('primary_keep', primary_keep)
    invoker.add_opt('secondary_get', secondary_get)
    invoker.add_opt('secondary_days', secondary_days)
    invoker.add_opt('secondary_when', secondary_when)
    invoker.add_opt('secondary_id_col', secondary_id_col)
    invoker.add_opt('secondary_date_col', secondary_date_col)
    invoker.add_opt('secondary_id2_col', secondary_id2_col)
    invoker.add_opt('merge_results', merge_results)
    invoker.add_opt('keep_original', keep_original)
    invoker.add_arg('primary', primary)
    invoker.add_arg('secondary', secondary)

    params = LinkCmdParams(invoker.get_opt('verbose'),
                           invoker.get_opt('id_col'),
                           invoker.get_opt('date_col'),
                           invoker.get_opt('id2_col'),
                           invoker.get_opt('primary_keep'),
                           invoker.get_opt('secondary_get'),
                           invoker.get_opt('secondary_days'),
                           invoker.get_opt('secondary_when'),
                           invoker.get_opt('secondary_id_col'),
                           invoker.get_opt('secondary_date_col'),
                           invoker.get_opt('secondary_id2_col'),
                           invoker.get_opt('merge_results'),
                           invoker.get_opt('keep_original'),
                           invoker.get_arg('primary'),
                           invoker.get_arg('secondary'))

    query = LinkQuery(params.opts, params.args)
    query.execute_query()

    databook = query.get_results()
    databook.add_metadata_sheet(invoker.get_command_info())
    databook.add_metadata_sheet(invoker.get_system_info())
    databook.to_excel(invoker.results_file)

    format_basic(invoker.results_file)
    format_merged_results(invoker.results_file)
    format_dups(invoker.results_file)

    post_msg1 = ('\nNOTE: If you want to merge/filter fields from the linked data in the '
                 'above results file, perform the following steps:')

    post_msg2 = (f'\t1. Open the results file and go to the "{SHEETNAME_FIELDS_AVAILABLE}" tab')
    post_msg3 = (f'\t2. In the column "{COL_HEADER_FIELDS_SELECTED}", '
                 'mark an "x" in each field you want to merge/keep')
    post_msg4 = ('\t3. Save the file')
    post_msg5 = ('\t4. Execute this command: "macpie merge FILE", '
                 'where FILE is the file you just saved.')
    post_msg6 = ('\t5. A new results file will be created containing the merged results\n')

    invoker.add_post_message(post_msg1)
    invoker.add_post_message(post_msg2)
    invoker.add_post_message(post_msg3)
    invoker.add_post_message(post_msg4)
    invoker.add_post_message(post_msg5)
    invoker.add_post_message(post_msg6)


def format_dups(filepath):
    wb = pyxl.load_workbook(filepath)
    for ws in wb.worksheets:
        if ws.title.endswith(SHEETNAME_SUFFIX_DUPLICATES):
            click.echo(f"Highlighting duplicates for {ws.title}...")
            util.pyxl.ws_highlight_rows_with_col(ws, COL_HEADER_DUPLICATES)
    wb.save(filepath)


def format_merged_results(filepath):
    wb = pyxl.load_workbook(filepath)
    if SHEETNAME_MERGED_RESULTS in wb.sheetnames:
        ws = wb[SHEETNAME_MERGED_RESULTS]
        # forced to keep the index column due to bug, so might as well give it a good name
        ws['A2'].value = COL_HEADER_ROW_INDEX
        wb.save(filepath)


class LinkCmdParams(CmdParams):

    def __init__(self,
                 verbose,
                 id_col,
                 date_col,
                 id2_col,
                 primary_keep,
                 secondary_get,
                 secondary_days,
                 secondary_when,
                 secondary_id_col,
                 secondary_date_col,
                 secondary_id2_col,
                 merge_results,
                 keep_original,
                 primary,
                 secondary) -> None:

        _primary = io.validate_filepath(primary, allowed_file)

        if secondary:
            (_secondary_valid, _secondary_invalid) = io.validate_filepaths(secondary, allowed_file)

            for p in _secondary_invalid:
                click.echo(f"WARNING: Ignoring invalid file: {p}")

            if len(_secondary_valid) < 1:
                raise click.UsageError("ERROR: No valid files.")
            elif len(_secondary_valid) == 1:
                if _primary == _secondary_valid[0]:
                    raise click.UsageError(f"ERROR: Primary file is {_primary}. No secondary files to link to.")

        self._opts = {
            'verbose': verbose,
            'id_col': id_col,
            'date_col': date_col,
            'id2_col': id2_col,
            'primary_keep': primary_keep,
            'secondary_get': secondary_get,
            'secondary_days': secondary_days,
            'secondary_when': secondary_when,
            'secondary_id_col': secondary_id_col if secondary_id_col is not None else id_col,
            'secondary_date_col': secondary_date_col if secondary_date_col is not None else date_col,
            'secondary_id2_col': secondary_id2_col if secondary_id2_col is not None else id2_col,
            'merge_results': util.validators.validate_bool_kwarg(merge_results, 'merge_results'),
            'keep_original': keep_original
        }

        self._args = {
            'primary': _primary,
            'secondary': _secondary_valid
        }

    @property
    def opts(self):
        return self._opts

    @property
    def args(self):
        return self._args


class LinkQuery:
    """
    The Receiver classes contain some important business logic. They know how to
    perform all kinds of operations, associated with carrying out a request. In
    fact, any class may serve as a Receiver.
    """

    def __init__(self, opts, args):
        self.opts = opts
        self.args = args
        self.Q = Query()

        try:
            primary_do = LavaDataObject.from_file(
                args['primary'],
                args['primary'].stem,
                id_col=opts['id_col'],
                date_col=opts['date_col'],
                id2_col=opts['id2_col']
            )
        except errors.DataObjectIDColKeyError:
            click.echo('\nWARNING: ID Column Header (-i, --id-col) not specified '
                       'and default of "InstrID" not found in your PRIMARY file.')
            click.echo(f'         Creating one for you called "{COL_HEADER_LINK_ID}"\n')

            primary_do = LavaDataObject.from_file(
                args['primary'],
                args['primary'].stem,
                id_col=None,
                date_col=opts['date_col'],
                id2_col=opts['id2_col']
            )
            primary_do.rename_id_col(COL_HEADER_LINK_ID)

        self.Q.add_node(
            primary_do,
            name=primary_do.name,
            operation=partial(pandas.group_by_keep_one,
                              group_by_col=primary_do.id2_col,
                              date_col=primary_do.date_col,
                              keep=opts['primary_keep'],
                              id_col=primary_do.id_col,
                              drop_duplicates=False)
        )

        if args['secondary']:
            for sec in args['secondary']:
                try:
                    secondary_do = LavaDataObject.from_file(
                        sec,
                        sec.stem,
                        id_col=opts['secondary_id_col'],
                        date_col=opts['secondary_date_col'],
                        id2_col=opts['secondary_id2_col']
                    )

                    self.Q.add_node(secondary_do)

                    self.Q.add_edge(
                        primary_do,
                        secondary_do,
                        name=secondary_do.name,
                        operation=partial(pandas.date_proximity,
                                          id_left_on=primary_do.id2_col,
                                          id_right_on=secondary_do.id2_col,
                                          date_left_on=primary_do.date_col,
                                          date_right_on=secondary_do.date_col,
                                          get=opts['secondary_get'],
                                          when=opts['secondary_when'],
                                          days=opts['secondary_days'],
                                          left_link_id=primary_do.id_col)
                    )

                except errors.ParserError:
                    click.echo(f"WARNING: Ignoring un-importable file: {sec}")

    def execute_query(self):
        click.echo("Executing query...")
        self.Q.execute()

    def get_results(self):
        db = Databook()
        db_dups = Databook()

        primary_node = self.Q.get_root_node()
        primary_result = self.Q.get_node(primary_node, 'operation_result')

        fields_list = [(primary_node.name,
                        col,
                        'x' if col in (primary_node.id_col,
                                       primary_node.date_col,
                                       primary_node.id2_col) else None)
                       for col in primary_result.columns
                       if col not in [primary_node.id_col, primary_node.date_col, primary_node.id2_col]]
        fields_list_dups = []

        if self.opts['merge_results'] is False:
            # if not merging, create new sheet for each result
            primary_sheet = Datasheet(primary_node.name, primary_result)
            primary_sheet.add_suffix(SHEETNAME_SUFFIX_PRIMARY)
            db.add_sheet(primary_sheet)

            edges_with_operation_results = self.Q.get_all_edge_data('operation_result')
            for edge in edges_with_operation_results:
                operation_result = edge['operation_result']
                secondary_sheet = Datasheet(edge['name'], operation_result)
                secondary_sheet.add_suffix(SHEETNAME_SUFFIX_SECONDARY)
                if operation_result[COL_HEADER_DUPLICATES].any():
                    secondary_sheet.add_suffix(SHEETNAME_SUFFIX_DUPLICATES)
                db.add_sheet(secondary_sheet)
                fields_list.extend([(edge['name'], col, None)
                                    for col in edge['operation_result'].columns
                                    if not col.endswith('_link')])
        else:
            # logic for creating a merged results sheet
            primary_result.columns = pd.MultiIndex.from_product([[primary_node.name], primary_result.columns])
            final_result = primary_result

            for left_node, right_node, operation_result in self.Q.g.edges.data('operation_result'):
                if operation_result is not None:
                    if operation_result[COL_HEADER_DUPLICATES].any():
                        # if there are dups
                        dup_sheet = Datasheet(self.Q.g.edges[left_node, right_node]['name'], operation_result)
                        dup_sheet.add_suffix(SHEETNAME_SUFFIX_SECONDARY)
                        dup_sheet.add_suffix(SHEETNAME_SUFFIX_DUPLICATES)
                        # put dups as separate worksheets after the results worksheet
                        db_dups.add_sheet(dup_sheet)
                        fields_list_dups.extend([(self.Q.g.edges[left_node, right_node]['name'], col, None)
                                                 for col in operation_result.columns
                                                 if not col.endswith('_link')])
                    else:
                        fields_list.extend([(self.Q.g.edges[left_node, right_node]['name'], col, None)
                                            for col in operation_result.columns
                                            if not col.endswith('_link')])

                        # merge all secondary results that do not have any duplicates
                        final_result = final_result.mac.merge(
                            operation_result,
                            left_on=[
                                (primary_node.name, primary_node.id2_col),
                                (primary_node.name, primary_node.date_col),
                                (primary_node.name, primary_node.id_col)
                            ],
                            # mac.date_proximity function adds the '_link' suffix
                            right_on=[col + '_link' for col in [primary_node.id2_col,
                                                                primary_node.date_col,
                                                                primary_node.id_col]],
                            add_indexes=(None, right_node.name)
                        )

            final_result.index = np.arange(1, len(final_result) + 1)
            merged_results_sheet = Datasheet(SHEETNAME_MERGED_RESULTS, final_result, display_index=True)
            db.add_sheet(merged_results_sheet)
            db.add_book(db_dups)

        # create meta sheets
        db.add_sheet(LinkQuery.create_fields_available_sheet(fields_list, fields_list_dups))
        db.add_metadata_sheet(self.create_merge_info_sheet())

        # create log sheets
        db.add_metadata_sheet(self.Q.get_log_dataobjects())
        db.add_metadata_sheet(self.Q.get_log_operations())

        return db

    @staticmethod
    def create_fields_available_sheet(fields_list, fields_list_dups=None):
        if fields_list_dups is not None:
            fields_list = fields_list + fields_list_dups
        fields_list_cols = ['DataObject', 'Field', COL_HEADER_FIELDS_SELECTED]
        fields_list_df = pd.DataFrame(data=fields_list, columns=fields_list_cols)
        return Datasheet(SHEETNAME_FIELDS_AVAILABLE, fields_list_df)

    def create_merge_info_sheet(self):
        # create merge info sheet
        primary_node = self.Q.get_root_node()
        secondary_do_names = [o['right_operand'] for o in self.Q.log_operations if o['operation'] == 'date_proximity']
        merge_info_data = [
            ('primary', primary_node.to_dict()),
            ('secondary', secondary_do_names),
            ('merged', self.opts['merge_results'])
        ]
        merge_info_df = pd.DataFrame(data=merge_info_data, columns=['param_name', 'param_value'])
        return Datasheet(SHEETNAME_MERGE_INFO, merge_info_df)
