from pathlib import Path
from shutil import copy

from click.testing import CliRunner
import openpyxl as pyxl
import pytest

from macpie._config import get_option
from macpie.io.excel import MACPieExcelWriter
from macpie.testing import assert_dfs_equal
from macpie.util.datasetfields import DatasetFields

from macpie.cli.main import main


current_dir = Path("tests/cli/merge/").resolve()

# output_dir = current_dir
output_dir = None


def create_available_fields(filepath):
    # manually create some fields to merge and test
    # obviously these fields should be available from the link results file
    available_fields_rows = [
        ('full', 'Col1'),
        ('full', 'Col2'),
        ('full', 'Col3'),
        ('instr2_all', 'Col1'),
        ('instr2_all', 'Col2'),
        ('instr2_all', 'Col3'),
        ('instr3_all', 'Col1'),
        ('instr3_all', 'Col2'),
        ('instr3_all', 'Col3')
    ]

    available_fields = DatasetFields(*available_fields_rows, title=get_option("sheet.name.available_fields"))
    available_fields.append_col_fill("x", header=get_option("column.to_merge"))

    wb = pyxl.load_workbook(filepath)
    del wb[get_option("sheet.name.available_fields")]
    wb.save(filepath)

    with MACPieExcelWriter(filepath, mode='a') as writer:
        available_fields.to_excel(writer)


@pytest.mark.slow
def test_full_no_merge(cli_link_full_no_merge, helpers, tmp_path):
    expected_result = helpers.read_merged_results(current_dir / "full_expected_results.xlsx")

    copied_file = Path(copy(cli_link_full_no_merge, tmp_path))
    create_available_fields(copied_file)

    runner = CliRunner()
    cli_args = ['merge', str(copied_file.resolve())]

    with runner.isolated_filesystem():
        results = runner.invoke(main, cli_args)
        assert results.exit_code == 0

        # get the results file
        results_path = next(Path(".").glob('**/result*xlsx'))

        # copy file to current dir if you want to debug more
        if output_dir is not None:
            copy(results_path, current_dir)

        results_wb = pyxl.load_workbook(results_path)
        # expected_results_wb = pyxl.load_workbook(current_dir / "full_expected_results.xlsx")

        expected_sheetnames = [
            'MERGED_RESULTS',
            'instr2_all_DUPS',
            'instr3_all_DUPS',
            get_option("sheet.name.available_fields"),
            get_option("sheet.name.collection_info")
        ]

        assert all(sheetname in results_wb.sheetnames for sheetname in expected_sheetnames)

        results = helpers.read_merged_results(results_path)
        assert_dfs_equal(results, expected_result, output_dir=output_dir)
