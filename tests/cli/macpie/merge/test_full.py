from pathlib import Path
from shutil import copy

from click.testing import CliRunner
import openpyxl as pyxl
import pandas as pd
import pytest

import macpie as mp
from macpie.io.excel import (
    DATASETS_SHEET_NAME,
    COLLECTION_SHEET_NAME,
)
from macpie.testing import DebugDir
from macpie.cli.macpie.main import main


THIS_DIR = Path(__file__).parent.absolute()


def create_available_fields(filepath):
    # manually create some fields to merge and test
    # obviously these fields should be available from the link results file
    available_fields_rows = [
        ("full", "Col1"),
        ("full", "Col2"),
        ("full", "Col3"),
        ("instr2_all", "Col1"),
        ("instr2_all", "Col2"),
        ("instr2_all", "Col3"),
        ("instr3_all", "Col1"),
        ("instr3_all", "Col2"),
        ("instr3_all", "Col3"),
    ]

    available_fields = mp.DatasetFields(
        *available_fields_rows, title=mp.MergeableAnchoredList.available_fields_sheetname
    )
    available_fields.append_col_fill("x", header=mp.MergeableAnchoredList.to_merge_column_name)

    wb = pyxl.load_workbook(filepath)
    del wb[mp.MergeableAnchoredList.available_fields_sheetname]
    wb.save(filepath)

    with mp.MACPieExcelWriter(filepath, mode="a", engine="mp_openpyxl") as writer:
        available_fields.to_excel(writer)


@pytest.mark.slow
def test_full_no_merge(cli_link_full_no_merge, tmp_path, debugdir):
    cli_link_full_no_merge_copy = Path(copy(cli_link_full_no_merge, tmp_path))

    expected_result = pd.read_excel(
        THIS_DIR / "full_expected_results.xlsx",
        sheet_name=mp.MergeableAnchoredList.merged_dsetname,
        header=[0, 1],
        index_col=None,
    )

    create_available_fields(cli_link_full_no_merge_copy)

    runner = CliRunner()
    cli_args = ["merge", str(cli_link_full_no_merge_copy.resolve())]

    with runner.isolated_filesystem(temp_dir=tmp_path):
        results = runner.invoke(main, cli_args)
        assert results.exit_code == 0

        # get the results file
        results_path = next(Path(".").glob("**/*.xlsx"))

        # copy file to current dir if you want to debug more
        if debugdir:
            with DebugDir(debugdir):
                copy(results_path, debugdir)

        results_wb = pyxl.load_workbook(results_path)
        # expected_results_wb = pyxl.load_workbook(current_dir / "full_expected_results.xlsx")

        expected_sheetnames = [
            "instr2_all_DUPS",
            "instr3_all_DUPS",
            COLLECTION_SHEET_NAME,
            DATASETS_SHEET_NAME,
            mp.MergeableAnchoredList.available_fields_sheetname,
            mp.MergeableAnchoredList.merged_dsetname,
        ]

        assert all(sheetname in results_wb.sheetnames for sheetname in expected_sheetnames)

        results = pd.read_excel(
            results_path,
            sheet_name=mp.MergeableAnchoredList.merged_dsetname,
            header=[0, 1],
            index_col=None,
        )

        pd.testing.assert_frame_equal(results, expected_result)
